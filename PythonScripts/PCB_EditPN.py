import re
import os
import xlrd
import pathlib
import logging
import openpyxl
import pythoncom
from win32com import client

logging.basicConfig(level=logging.DEBUG, filename="PCB_edit_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")


def ReadBom(customer_bom, designator, pn, start_row, end_row, delimiter, separator, bot_pcb, top_pcb):
    try:
        file_extension = pathlib.Path(customer_bom).suffix
        logging.info("Reading Customer BOM Excel...")
        bom_data = []
        sep_data = []
        count = 0
        bom_col_des = ord(designator) - 65
        bom_col_pn = ord(pn) - 65
        if file_extension == ".xls" or file_extension == ".XLS":
            wb_bom = xlrd.open_workbook(customer_bom)
            ws_bom = wb_bom.sheet_by_index(0)
            if separator is not None:
                for row in range(start_row - 1, end_row):
                    var = ws_bom.row_values(row)
                    x = var[bom_col_des]
                    if bool(x):
                        if delimiter is not None:
                            x = x.replace(" ", "").split(delimiter)
                            x = list(filter(None, x))
                            for item in x:
                                if separator in item:
                                    sep_data.append([item, count, 0])
                    count += 1
                    bom_data.append([x, var[bom_col_pn], 0])
                return [sep_data, bom_data, file_extension, bom_col_des], None
            else:
                for row in range(start_row - 1, end_row):
                    var = ws_bom.row_values(row)
                    x = var[bom_col_des]
                    if bool(x):
                        if delimiter is not None:
                            x = x.replace(" ", "").split(delimiter)
                            x = list(filter(None, x))
                    bom_data.append([x, var[bom_col_pn], 0])
                logging.info("Finished reading")
                result, error = modify(customer_bom, bot_pcb, top_pcb, bom_data)
                if result is not True:
                    return False, error
                return True, None
        else:
            wb_bom = openpyxl.load_workbook(customer_bom)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            if separator is not None:
                for row in bom_rows[int(start_row) - 1:int(end_row)]:
                    x = row[bom_col_des].value
                    if bool(x):
                        if delimiter is not None:
                            x = x.replace(" ", "").split(delimiter)
                            x = list(filter(None, x))
                            for item in x:
                                if separator in item:
                                    sep_data.append([item, count, 0])
                    count += 1
                    bom_data.append([x, row[bom_col_pn].value, 0])
                wb_bom.close()
                return [sep_data, bom_data, file_extension, bom_col_des], None
            else:
                for row in bom_rows[int(start_row) - 1:int(end_row)]:
                    x = row[bom_col_des].value
                    if bool(x):
                        if delimiter is not None:
                            x = x.replace(" ", "").split(delimiter)
                            x = list(filter(None, x))
                    bom_data.append([x, row[bom_col_pn].value, 0])
                wb_bom.close()
                logging.info("Finished reading")
                result, error = modify(customer_bom, bot_pcb, top_pcb, bom_data)
                if result is not True:
                    return False, error
                return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 85")
        logging.error("Error while reading Customer BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Customer BOM File!")
        return False, f"Error while reading Customer BOM File!\n{e.__class__}\n{e}"


def CustBomInfo(sep_detail, bom_data, customer_bom, bot_pcb, top_pcb):
    for item in sep_detail:
        x = bom_data[item[1]][0].index(item[0])
        bom_data[item[1]][0].pop(x)
        bom_data[item[1]][0].extend(item[3])
    logging.info("Finished reading")
    result, error = modify(customer_bom, bot_pcb, top_pcb, bom_data)
    if result is not True:
        return False, error
    return True, None


def modify(bomfile, bot_file, top_file, bom_data):
    try:
        logging.info("Reading first pcb file...")
        botrefdes = []
        bot_textfile = open(bot_file, 'r')
        botfiledata = bot_textfile.readlines()
        for line in botfiledata:
            match = re.match("^F9\s", line)
            if match:
                botrefdes.append([line.strip().split(" ").pop()])
        bot_textfile.close()
        bot_bom_data = bom_data.copy()
        if bool(top_file):
            logging.info("Reading second pcb file...")
            toprefdes = []
            top_textfile = open(top_file, 'r')
            topfiledata = top_textfile.readlines()
            for line in topfiledata:
                match = re.match("^F9\s", line)
                if match:
                    toprefdes.append([line.strip().split(" ").pop()])
            top_textfile.close()
            top_bom_data = bom_data.copy()

        logging.info("Finished reading pcb files!")

    except Exception as e:
        logging.error(f"{e.__class__} from line 80")
        logging.error("Error while reading .pcb files!")
        logging.error(f"{e}")
        print(e, "\n Error while reading .pcb files!")
        return False, f"Error while eading .pcb files!\n{e.__class__}\n{e}"

    try:
        logging.info("Mapping pcb RefDes to BOM excel...")
        pointer = 0
        for value in botrefdes:
            flag = 0
            for ref in bot_bom_data:
                if bool(ref[0]) and value[0] in ref[0] and ref[1] is not None:
                    botrefdes[pointer].append(ref[1])
                    ref[2] = 1
                    flag = 1
                    break
            if flag == 0:
                botrefdes[pointer].append("Not found")
            pointer += 1
        if bool(top_file):
            logging.info("Mapping pcb RefDes to BOM excel...")
            pointer = 0
            for value in toprefdes:
                flag = 0
                for ref in top_bom_data:
                    if bool(ref[0]) and value[0] in ref[0] and ref[1] is not None:
                        toprefdes[pointer].append(ref[1])
                        ref[2] = 1
                        flag = 1
                        break
                if flag == 0:
                    toprefdes[pointer].append("Not found")
                pointer += 1

        logging.info("Finished mapping!")

    except Exception as e:
        logging.error(f"{e.__class__} from line 192")
        logging.error("Error while mapping designators from .pcb files to excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from .pcb files to excel file!")
        return False, f"Error while mapping designators from .pcb files to excel file!\n{e.__class__}\n{e}"

    try:
        logging.info("Writing modified pcb file...")
        pointer = 0
        line_pointer = 0
        for line in botfiledata:
            match = re.match("^F8\s", line)
            if match:
                if botrefdes[pointer][1] != "Not found" and botrefdes[pointer][1] is not None:
                    x = line.strip().split(" ")
                    x = x[0:7]
                    x.append(botrefdes[pointer][1])
                    x = " ".join(x)
                    botfiledata[line_pointer] = x + "\n"
                pointer += 1
            line_pointer += 1

            var = bot_textfile.name.split("/")
            new_file_name = var.pop().replace(".pcb", "_modified.pcb")
            new_file_name = new_file_name.replace(".PCB", "_modified.PCB")
            var.append(new_file_name)
            new_file = "/".join(var)

            with open(new_file, "w") as f:
                for item in botfiledata:
                    f.write("%s" % item)
            f.close()

        if bool(top_file):
            pointer = 0
            line_pointer = 0
            for line in topfiledata:
                match = re.match("^F8\s", line)
                if match:
                    if toprefdes[pointer][1] != "Not found" and toprefdes[pointer][1] is not None:
                        x = line.strip().split(" ")
                        x = x[0:7]
                        x.append(toprefdes[pointer][1])
                        x = " ".join(x)
                        topfiledata[line_pointer] = x + "\n"
                    pointer += 1
                line_pointer += 1

            var = top_textfile.name.split("/")
            new_file_name = var.pop().replace(".pcb", "_modified.pcb")
            new_file_name = new_file_name.replace(".PCB", "_modified.PCB")
            var.append(new_file_name)
            new_file = "/".join(var)

            with open(new_file, "w") as f:
                for item in topfiledata:
                    f.write("%s" % item)
            f.close()
            logging.info("Finished writing!")

        logging.info("Finished writing!")
        if bool(top_file):
            result, error = WriteBom(bomfile, bot_bom_data, top_bom_data, botrefdes, toprefdes)
            if result is not True:
                return False, error
            return True, None
        else:
            result, error = SingleBom(bomfile, bot_bom_data, botrefdes)
            if result is not True:
                return False, error
            return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 272")
        logging.error("Error while creating modified .pcb files!")
        logging.error(f"{e}")
        print(e, "\n Error while creating modified .pcb files!")
        return False, f"Error while creating modified .pcb files!\n{e.__class__}\n{e}"


def WriteBom(bomfile, bot_bom_data, top_bom_data, botrefdes, toprefdes):
    try:
        logging.info("Writing to BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            pythoncom.CoInitialize()
            xlApp = client.Dispatch("Excel.Application")
            fileDir = os.path.dirname(os.path.realpath('__file__'))
            filename = os.path.join(fileDir, f'{bomfile}')
            wkbk = xlApp.Workbooks.open(filename)
            wksht = wkbk.Worksheets(1)
            col = 5
            count = 0
            missing = []
            row = int(wksht.UsedRange.Rows.Count) + 3
            wksht.Cells(row, 2).Value = "Missing values in Bottom pcb:"
            for item in bot_bom_data:
                if item[2] == 0:
                    if top_bom_data[count][2] == 0:
                        missing.append(item[0])
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col + 1).Value = item[1]
                    row += 1
                count += 1

            count = 0
            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in Top pcb:"
            for item in top_bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col + 1).Value = item[1]
                    row += 1
                count += 1

            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in both:"
            for item in missing:
                if bool(item):
                    val = ", ".join(item)
                    wksht.Cells(row, col).Value = val
                    row += 1

            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in BOM:"
            for item in botrefdes:
                if item[1] == "Not found":
                    # wksht.Cells(row, col).Font.ColorIndex = 3
                    wksht.Cells(row, col).Value = item[0]
                    row += 1
            for item in toprefdes:
                if item[1] == "Not found":
                    # wksht.Cells(row, col).Font.ColorIndex = 3
                    wksht.Cells(row, col).Value = item[0]
                    row += 1

            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()

        else:
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            col = 5
            count = 0
            missing = []
            row = int(ws_bom.max_row) + 3
            ws_bom.cell(row=row, column=2).value = "Missing values in Bottom pcb:"
            for item in bot_bom_data:
                if item[2] == 0:
                    if top_bom_data[count][2] == 0:
                        missing.append(item[0])
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col + 1).value = item[1]
                    row += 1
                count += 1

            count = 0
            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in Top pcb:"
            for item in top_bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col + 1).value = item[1]
                    row += 1
                count += 1

            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in both:"
            for item in missing:
                if bool(item):
                    val = ", ".join(item)
                    ws_bom.cell(row=row, column=col).value = val
                    row += 1

            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in BOM:"
            for item in botrefdes:
                if item[1] == "Not found":
                    # ws_bom.cell(row=row, column=col).font = Font(color="00FF0000")
                    ws_bom.cell(row=row, column=col).value = item[0]
                    row += 1
            for item in toprefdes:
                if item[1] == "Not found":
                    # ws_bom.cell(row=row, column=col).font = Font(color="00FF0000")
                    ws_bom.cell(row=row, column=col).value = item[0]
                    row += 1

            wb_bom.save(bomfile)
        logging.info("Finished writing!")
        return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 309")
        logging.error("Error while writing to BOM excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing to BOM excel file!")
        return False, f"Error while writing to BOM excel file!\n{e.__class__}\n{e}"


def SingleBom(bomfile, bom_data, pcbrefdes):
    try:
        logging.info("Writing to BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            pythoncom.CoInitialize()
            xlApp = client.Dispatch("Excel.Application")
            fileDir = os.path.dirname(os.path.realpath('__file__'))
            filename = os.path.join(fileDir, f'{bomfile}')
            wkbk = xlApp.Workbooks.open(filename)
            wksht = wkbk.Worksheets(1)
            col = 5
            count = 0
            row = int(wksht.UsedRange.Rows.Count) + 3
            wksht.Cells(row, 2).Value = "Missing values in PCB File:"
            for item in bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col+1).Value = item[1]
                    row += 1
                count += 1

            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in BOM:"
            for item in pcbrefdes:
                if item[1] == "Not found":
                    # wksht.Cells(row, col).Font.Color = 30
                    wksht.Cells(row, col).Value = item[0]
                    row += 1

            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()

        else:
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            col = 5
            count = 0
            row = int(ws_bom.max_row) + 3
            ws_bom.cell(row=row, column=2).value = "Missing values in PCB File:"
            for item in bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col+1).value = item[1]
                    row += 1
                count += 1

            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in BOM:"
            for item in pcbrefdes:
                if item[1] == "Not found":
                    # ws_bom.cell(row=row, column=col).font = Font(color="00FF0000")
                    ws_bom.cell(row=row, column=col).value = item[0]
                    row += 1
            wb_bom.save(bomfile)
        logging.info("Finished writing!")
        return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 239")
        logging.error("Error while writing to BOM excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing to BOM excel file!")
        return False, f"Error while writing to BOM excel file!\n{e.__class__}\n{e}"


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")
    logging.info("Successfully Executed!!!\n\n")
    print("Successfully Executed!!!")
