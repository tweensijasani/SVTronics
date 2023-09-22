import re
import os
import xlrd
import pathlib
import logging
import openpyxl
import datetime
import pythoncom
import configparser
from win32com import client
from openpyxl.styles import Font

logging.basicConfig(level=logging.DEBUG, filename="Excel_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")


def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def CheckBom(customer_bom, manex_bom, designator, quantity, start_row, end_row, delimiter, separator):
    try:
        result, error = CheckManexBom(manex_bom)
        if not bool(result):
            return result, error
        return CheckCustBom(customer_bom, manex_bom, designator, quantity, start_row, end_row, delimiter, separator, result)

    except Exception as e:
        logging.error(f"{e.__class__} from line 168")
        logging.error("Error while checking BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while checking BOM!")
        return False, f"Error while checking BOM\n{e.__class__}\n{e}"


def CheckManexBom(manex_bom):
    try:
        logging.info("Reading setup.ini file...")
        config = configparser.ConfigParser()
        config.read('setup.ini')

        manex_itemno = config['INITIALIZATION']['item_no']
        manex_designator = config['INITIALIZATION']['RefDes']
        manex_quantity = config['INITIALIZATION']['Quantity']
        manex_partno = config['INITIALIZATION']['PartNumber']
        manex_start_row = int(config['INITIALIZATION']['StartRow'])
        manex_delimiter = config['INITIALIZATION']['Delimiter']
        manex_separator = config['INITIALIZATION']['Separator']
        logging.info("Info populated")

    except Exception as e:
        logging.error(f"{e.__class__} from line 168")
        logging.error("Error while getting Manex BOM Detail Inputs!")
        logging.error(f"{e}")
        print(e, "\n Error while getting Manex BOM Detail Inputs!")
        return False, f"Error while setup.ini data\n{e.__class__}\n{e}"

    try:
        logging.info("Reading Manex BOM Excel...")
        wb_manex = openpyxl.load_workbook(manex_bom)
        ws_manex = wb_manex.worksheets[0]
        manex_end_row = int(ws_manex.max_row)
        manex_rows = list(ws_manex.rows)
        header = []
        for values in manex_rows[0]:
            header.append(values.value)
        try:
            manex_col_itemno = header.index(manex_itemno)
            manex_col_des = header.index(manex_designator)
            manex_col_qty = header.index(manex_quantity)
            manex_col_partno = header.index(manex_partno)
        except Exception as e:
            logging.error(f"{e.__class__} from line 214")
            logging.error("Can't locate item_no/RefDesg/QtEach/PART_NO in Manex BOM!!")
            logging.error(f"{e}")
            return False, f"Error while reading Manex BOM\n{e.__class__}\n{e}"

        error_msg = []
        manex_data = []
        for row in manex_rows[int(manex_start_row)-1:int(manex_end_row)]:
            y1 = row[manex_col_des].value
            y2 = str(row[manex_col_qty].value)
            y3 = row[manex_col_partno].value
            y4 = row[manex_col_itemno].value
            if bool(y1):
                if manex_delimiter is not None:
                    y1 = (row[manex_col_des].value).replace(" ", "").split(manex_delimiter)
                    y1 = list(filter(None, y1))
            else:
                error_msg.append(f"Missing designator for item no {y4}")
                logging.error(f"Missing designator for item no {y4}")
            if not bool(y2):
                error_msg.append(f"Missing quantity for item no {y4}")
                logging.error(f"Missing quantity for item no {y4}")
            elif not y2.isnumeric():
                error_msg.append(f"Non-integer quantity for item no {y4}")
                logging.error(f"Non-integer quantity for item no {y4}")
            elif int(y2) == 0:
                error_msg.append(f"Zero quantity for item no {y4}")
                logging.error(f"Zero quantity for item no {y4}")
            if not bool(y3):
                error_msg.append(f"Missing part-no for item no {y4}")
                logging.error(f"Missing part-no for item no {y4}")
            manex_data.append([y1, y2, y3])
        logging.info("Finished reading")
        wb_manex.close()
        if bool(error_msg):
            logging.error("Terminated!! Manex BOM not clean!!")
            return False, ["Manex BOM is not clean!!", error_msg]
        return {'manex_data': manex_data, 'manex_start_row': manex_start_row, 'manex_end_row': manex_end_row, 'manex_col_partno': manex_col_partno, 'manex_separator': manex_separator}, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 190")
        logging.error("Error while reading Manex BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Manex BOM File!")
        return False, f"Error while reading Manex BOM\n{e.__class__}\n{e}"


def CheckCustBom(customer_bom, manex_bom, designator, quantity, start_row, end_row, delimiter, separator, manex_dict):
    try:
        file_extension = pathlib.Path(customer_bom).suffix
        logging.info("Reading Customer BOM Excel...")
        bom_data = []
        error_msg = []
        itemno = 1
        bom_col_des = ord(designator) - 65
        bom_col_qty = ord(quantity) - 65
        if file_extension == ".xls" or file_extension == ".XLS":
            wb_bom = xlrd.open_workbook(customer_bom)
            ws_bom = wb_bom.sheet_by_index(0)
            for row in range(start_row-1, end_row):
                var = ws_bom.row_values(row)
                x1 = var[bom_col_des]
                x2 = str(var[bom_col_qty])
                if bool(x1):
                    x1 = x1.replace(" ", "").split(delimiter)
                    x1 = list(filter(None, x1))
                else:
                    error_msg.append(f"Missing designator for item no {itemno} or line no {row + 1}")
                    logging.error(f"Missing designator for item no {itemno} or line no {row + 1}")
                if not bool(x2):
                    error_msg.append(f"Missing quantity for item no {itemno} or line no {row + 1}")
                    logging.error(f"Missing quantity for item no {itemno} or line no {row + 1}")
                elif not x2.isnumeric():
                    if isfloat(x2):
                        if not float(x2).is_integer():
                            error_msg.append(f"Non-integer quantity for item no {itemno} or line no {row + 1}")
                            logging.error(f"Non-integer quantity for item no {itemno} or line no {row + 1}")
                        else:
                            x2 = int(float(x2))
                    else:
                        error_msg.append(f"Non-integer quantity for item no {itemno} or line no {row + 1}")
                        logging.error(f"Non-integer quantity for item no {itemno} or line no {row + 1}")
                bom_data.append([x1, x2, 1, row])
                itemno += 1
        else:
            wb_bom = openpyxl.load_workbook(customer_bom)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            count = start_row
            for row in bom_rows[int(start_row)-1:int(end_row)]:
                x1 = row[bom_col_des].value
                x2 = str(row[bom_col_qty].value)
                if bool(x1):
                    x1 = x1.replace(" ", "").split(delimiter)
                    x1 = list(filter(None, x1))
                else:
                    error_msg.append(f"Missing designator for item no {itemno} or line no {count}")
                    logging.error(f"Missing designator for item no {itemno} or line no {count}")
                if not bool(x2):
                    error_msg.append(f"Missing quantity for item no {itemno} or line no {count}")
                    logging.error(f"Missing quantity for item no {itemno} or line no {count}")
                elif not x2.isnumeric():
                    if isfloat(x2):
                        if not float(x2).is_integer():
                            error_msg.append(f"Non-integer quantity for item no {itemno} or line no {count}")
                            logging.error(f"Non-integer quantity for item no {itemno} or line no {count}")
                        else:
                            x2 = int(float(x2))
                    else:
                        error_msg.append(f"Non-integer quantity for item no {itemno} or line no {count}")
                        logging.error(f"Non-integer quantity for item no {itemno} or line no {count}")
                bom_data.append([x1, x2, 1, count])
                itemno += 1
                count += 1
            wb_bom.close()

        if bool(error_msg):
            logging.error("Terminated!! Customer BOM not clean!!")
            return False, ["Customer BOM is not clean!!", error_msg]

        return ReadCustBom(customer_bom, manex_bom, start_row, end_row, separator, bom_data, file_extension, bom_col_des, manex_dict)

    except Exception as e:
        logging.error(f"{e.__class__} from line 85")
        logging.error("Error while reading Customer BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Customer BOM File!")
        return False, f"Error while reading Customer BOM\n{e.__class__}\n{e}"


def ReadCustBom(customer_bom, manex_bom, start_row, end_row, separator, bom_data, file_extension, bom_col_des, manex_dict):

    try:
        sep_data = []
        sep_position = []
        count = 0
        if separator is not None:
            for data in bom_data:
                for item in data[0]:
                    if separator in item:
                        sep_data.append([item, count, 0])
                        sep_position.append([data[3], bom_data.index(data)])
                if len(data[0]) != int(data[1]):
                    qty = False
                    data[2] = qty
                data.pop()
                count += 1
            return [sep_data, bom_data, file_extension, bom_col_des, sep_position, manex_dict], None
        else:
            for data in bom_data:
                if len(data[0]) != int(data[1]):
                    qty = False
                    data[2] = qty
                data.pop()
            result, error = MapDes(customer_bom, manex_bom, bom_data, file_extension, start_row, end_row, bom_col_des, {}, [], manex_dict)
            if result is not True:
                return False, error
            return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 85")
        logging.error("Error while interpreting Customer BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while interpreting Customer BOM!")
        return False, f"Error while interpreting Customer BOM\n{e.__class__}\n{e}"


def CustBomInfo(sep_detail, bom_data, customer_bom, manex_bom, file_extension, start_row, end_row, bom_col_des, sep_dict, sep_position, manex_dict):
    for item in sep_detail:
        if item[2] == 1:
            x = bom_data[item[1]][0].index(item[0])
            bom_data[item[1]][0].pop(x)
            bom_data[item[1]][0].extend(item[3])
            if len(bom_data[item[1]][0]) != int(bom_data[item[1]][1]):
                bom_data[item[1]][2] = False
            else:
                bom_data[item[1]][2] = True
    logging.info("Finished reading")
    result, error = MapDes(customer_bom, manex_bom, bom_data, file_extension, start_row, end_row, bom_col_des, sep_dict, sep_position, manex_dict)
    if result is not True:
        return False, error
    return True, None


def ManexInfo(sep_dict, manex_dict):
    try:
        for data in manex_dict['manex_data']:
            y = data[0]
            rem = []
            new = []
            for item in y:
                if manex_dict['manex_separator'] in item:
                    if item in sep_dict:
                        rem.append(item)
                        new.extend(sep_dict[item])
                    else:
                        res = []
                        stry = item.split(manex_dict['manex_separator'])
                        str1 = stry[0]
                        str2 = stry[1]
                        base = ""
                        for i in range(min(len(str1), len(str2))):
                            if str1[i] == str2[i]:
                                base = f"{base}{str1[i]}"
                            else:
                                break
                        if base == '':
                            break
                        count1 = 0
                        count2 = 0
                        for i in range(len(base) - 1):
                            if str1[i] == base[i]:
                                count1 += 1
                            if str2[i] == base[i]:
                                count2 += 1
                        str1 = str1[count1 + 1:]
                        str2 = str2[count2 + 1:]
                        my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                        my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                        if len(my_list1) > 1:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                        res.append(f"{base}{chr(i)}{j}")
                            else:
                                for i in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                    for j in range(ord(my_list1[1]), ord(my_list2[1]) + 1):
                                        res.append(f"{base}{i}{chr(j)}")
                        else:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    res.append(f"{base}{chr(i)}")
                            else:
                                for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                    res.append(f"{base}{j}")
                        rem.append(item)
                        new.extend(res)
            if bool(rem):
                for obj in rem:
                    pointer = y.index(obj)
                    y.pop(pointer)
                y.extend(new)
            data[0] = y
        return manex_dict, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 190")
        logging.error("Error while reading Manex BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Manex BOM File!")
        return False, f"Error while reading Manex BOM\n{e.__class__}\n{e}"


def MapDes(customer_bom, manex_bom, bom_data, file_extension, start_row, end_row, bom_col_des, sep_dict, sep_position, manex_dict):

    result, error = ManexInfo(sep_dict, manex_dict)
    if not bool(result):
        return False, error
    manex_data = result['manex_data']
    manex_start_row = result['manex_start_row']
    manex_end_row = result['manex_end_row']
    manex_col_partno = result['manex_col_partno']

    try:
        logging.info("Mapping designators from Manex BOM to Customer BOM...")
        manex_pn = []
        duplicate = []
        # pcb = []
        for item in bom_data:
            if bool(item[0]):
                flag = 0
                pn = []
                for obj in manex_data:
                    if bool(obj[0]):
                        # if set("PCB").issubset(set(obj[0][0])):
                        #     if not bool(pcb):
                        #         pcb.append(obj[2])
                        if set(item[0]).issubset(set(obj[0])) or set(obj[0]).issubset(set(item[0])):
                            if obj[1] != 0:
                                if flag == 0:
                                    manex_pn.append(obj[2])
                                if obj[2] not in pn:
                                    pn.append(obj[2])
                                    flag += 1
                if flag == 0:
                    for obj in manex_data:
                        if bool(obj[0]):
                            if len(obj[0]) > 2:
                                rem = obj[0].pop()
                                if set(item[0]).issubset(set(obj[0])) or set(obj[0]).issubset(set(item[0])) or any(value in obj[0] for value in item[0]):
                                    if obj[1] != 0:
                                        if flag == 0:
                                            manex_pn.append(obj[2])
                                        if obj[2] not in pn:
                                            pn.append(obj[2])
                                            flag += 1
                                obj[0].append(rem)
                if flag == 0:
                    manex_pn.append('Not in Manex')
                if flag > 1:
                    duplicate.extend(pn)
            else:
                manex_pn.append(None)
        logging.info("Finished mapping")
        cust, error_cust = WriteCustBom(customer_bom, file_extension, start_row, end_row, manex_pn, duplicate, bom_col_des, bom_data, sep_position)
        man, error_man = WriteManexBom(manex_bom, manex_start_row, manex_end_row, manex_col_partno, manex_pn, duplicate)
        if cust is not True:
            return False, error_cust
        if man is not True:
            return False, error_man
        return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 22")
        logging.error("Error while mapping designators from Manex BOM to Customer BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from Manex BOM to Customer BOM!")
        return False, f"Error while mapping designators from Manex BOM to Customer BOM\n{e.__class__}\n{e}"


def WriteCustBom(customer_bom, file_extension, bom_start_row, bom_end_row, manex_pn, duplicate, bom_col_des, bom_data, sep_position):

    try:
        logging.info("Writing to Customer Bom Excel...")
        # if file_extension == ".xls" or file_extension == ".XLS":
        pythoncom.CoInitialize()
        xlApp = client.Dispatch("Excel.Application")
        fileDir = os.path.dirname(os.path.realpath('__file__'))
        filename = os.path.join(fileDir, f'{customer_bom}')
        wkbk = xlApp.Workbooks.open(filename)
        wksht = wkbk.Worksheets(1)
        wksht.Columns("A").EntireColumn.Insert()
        j = 0
        for i in range(bom_start_row, bom_end_row+1):
            wksht.Cells(i, 1).Value = manex_pn[j]
            j += 1
        wksht.Columns("A").EntireColumn.Insert()
        wksht.Rows(bom_end_row + 1).EntireRow.Insert()
        wksht.Rows(bom_end_row + 1).EntireRow.Insert()
        wksht.Cells(bom_end_row + 1, 3).Interior.ColorIndex = 0
        wksht.Cells(bom_end_row + 2, 3).Interior.ColorIndex = 0
        string = f"Manex PN added on {datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S')}"
        wksht.Cells(bom_end_row + 2, 2).Value = string
        pointer = 0
        for i in range(bom_start_row, bom_end_row+1):
            if wksht.Cells(i, 2).Value == "Not in Manex":
                wksht.Cells(i, 1).Value = "Check"
                wksht.Cells(i, 1).Interior.ColorIndex = 6
            elif wksht.Cells(i, 2).Value in duplicate:
                wksht.Cells(i, 1).Value = "Duplicate RefDesgs in BOM"
                wksht.Cells(i, 1).Interior.ColorIndex = 8
            if not bool(bom_data[pointer][2]):
                wksht.Cells(i, 1).Value = "Quantity Column and RefDesg Count does not match"
            if not bool(bom_data[pointer][1]):
                for col in range(1, int(wksht.UsedRange.Columns.Count)):
                    wksht.Cells(i, col).Font.ColorIndex = 3
            pointer += 1

        if file_extension == ".xls" or file_extension == ".XLS":
            for value in sep_position:
                wksht.Cells(value[0] + 1, bom_col_des + 3).Value = ", ".join(bom_data[value[1]][0])
        else:
            for value in sep_position:
                wksht.Cells(value[0], bom_col_des + 3).Value = ", ".join(bom_data[value[1]][0])

        wkbk.Save()
        wkbk.Close(True)
        xlApp.Quit()

        logging.info("Finished writing")
        return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 220")
        logging.error("Error while writing Customer BOM excel!")
        logging.error(f"{e}")
        print(e, "\n Error while writing Customer BOM excel!")
        return False, f"Error while writing into Customer BOM\n{e.__class__}\n{e}"


def WriteManexBom(manex_bom, manex_start_row, manex_end_row, manex_col_partno, manex_pn, duplicate):

    try:
        logging.info("Writing to Manex BOM Excel...")
        wb_manex = openpyxl.load_workbook(manex_bom)
        ws_manex = wb_manex.worksheets[0]
        r = manex_start_row
        for rows in ws_manex.iter_rows(min_row=manex_start_row, max_row=manex_end_row, min_col=1, max_col=20):
            if ws_manex[f"{chr(manex_col_partno+65)}{str(r)}"].value not in manex_pn:
                for cell in rows:
                    cell.font = Font(color="00FF1414")
            if ws_manex[f"{chr(manex_col_partno+65)}{str(r)}"].value in duplicate:
                for cell in rows:
                    cell.font = Font(color="000096FF")
            r += 1
        logging.info("Finished writing")
        wb_manex.save(manex_bom)
        return True, None

    except Exception as e:
        logging.error(f"{e.__class__} from line 279")
        logging.error("Error while writing Manex file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing Manex file!")
        wb_manex.close()
        return False, f"Error while writing into Manex BOM\n{e.__class__}\n{e}"


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")
    logging.info("Successfully Executed!!!\n\n")
    print("Successfully Executed!!!")

