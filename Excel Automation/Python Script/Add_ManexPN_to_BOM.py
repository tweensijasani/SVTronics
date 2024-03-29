import re
import os
import sys
import xlrd
import pathlib
import logging
import openpyxl
import datetime
import configparser
import easygui.boxes
import tkinter as tk
from easygui import *
from win32com import client
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill, Font

logging.basicConfig(level=logging.DEBUG, filename="Excel_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")

root = tk.Tk()
root.withdraw()

easygui.boxes.global_state.PROPORTIONAL_FONT_FAMILY = ("MS", "Arial")
easygui.boxes.global_state.MONOSPACE_FONT_FAMILY = "Arial"
easygui.boxes.global_state.PROPORTIONAL_FONT_SIZE = 14
easygui.boxes.global_state.MONOSPACE_FONT_SIZE = 14
easygui.boxes.global_state.TEXT_ENTRY_FONT_SIZE = 14
easygui.boxes.global_state.prop_font_line_length = 30
easygui.boxes.global_state.fixw_font_line_length = 40


def getfiles():

    try:
        counter = 0
        Manex_File = filedialog.askopenfilename(title="Select Web Manex BOM", filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
        while Manex_File == '' and counter < 1:
            messagebox.showerror(title="File Error", message="Manex BOM Not Selected")
            logging.warning("Manex BOM Not Selected")
            Manex_File = filedialog.askopenfilename(title="Select Web Manex BOM",
                                                    filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
            counter += 1
        if Manex_File == '':
            messagebox.showerror(title="Invalid Input", message="Terminated: Manex BOM not selected!")
            logging.error("Terminated: Manex BOM not selected!")
            sys.exit(1)
        else:
            logging.info("Manex BOM Excel Selected")

        counter = 0
        Bom_File = filedialog.askopenfilename(title="Select Customer BOM", filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
        while Bom_File == '' and counter < 1:
            messagebox.showerror(title="File Error", message="Customer BOM Not Selected")
            logging.warning("Customer BOM File Not Selected")
            Bom_File = filedialog.askopenfilename(title="Select Customer BOM", filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
            counter += 1
        if Bom_File == '':
            messagebox.showerror(title="Invalid Input", message="Terminated: Customer BOM not selected!")
            logging.error("Terminated: Customer BOM not selected!")
            sys.exit(1)
        else:
            logging.info("Customer BOM Excel Selected")
        return Manex_File, Bom_File

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while fetching files!!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__}")
        logging.error("Error while fetching files!")
        logging.error(f"{e}")
        print(e, "\n Error while fetching files!")
        sys.exit(1)


def manex_metadata():
    try:
        logging.info("Reading setup.ini file...")
        config = configparser.ConfigParser()
        config.read('setup.ini')

        manex_itemno = config['INITIALIZATION']['item_no']
        manex_designator = config['INITIALIZATION']['RefDes']
        manex_quantity = config['INITIALIZATION']['Quantity']
        manex_partno = config['INITIALIZATION']['PartNumber']
        manex_start_row = int(config['INITIALIZATION']['StartRow'])
        manex_delimiter = config['INITIALIZATION']['Delimiter']
        manex_separator = config['INITIALIZATION']['Separator']

        manex_dict = {"manex_itemno": manex_itemno, "manex_designator": manex_designator, "manex_quantity": manex_quantity,
                      "manex_partno": manex_partno, "manex_start_row": manex_start_row, "manex_delimiter": manex_delimiter, "manex_separator": manex_separator}
        logging.info("Info populated")
        return manex_dict

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while reading setup.ini!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 98")
        logging.error("Error while reading setup.ini!")
        logging.error(f"{e}")
        print(e, "\n Error while reading setup.ini!")
        sys.exit(1)


def Manex_data(Manex_File, manex_dict):
    try:
        logging.info("Reading Manex BOM Excel...")
        wb_manex = openpyxl.load_workbook(Manex_File, data_only=True)
        ws_manex = wb_manex.worksheets[0]
        manex_end_row = int(ws_manex.max_row)
        manex_dict.update({"manex_end_row": manex_end_row})
        manex_rows = list(ws_manex.rows)
        header = []
        for values in manex_rows[0]:
            header.append(values.value)
        try:
            manex_col_itemno = header.index(manex_dict["manex_itemno"])
            manex_col_des = header.index(manex_dict["manex_designator"])
            manex_col_qty = header.index(manex_dict["manex_quantity"])
            manex_col_partno = header.index(manex_dict["manex_partno"])
            manex_dict.update({"manex_col_partno": manex_col_partno})
        except Exception as e:
            messagebox.showerror(title=f"{e.__class__}", message=f"Can't locate item_no/RefDesg/QtEach/PART_NO in Manex BOM!!\n\n{e.__class__}\n{e}")
            logging.error(f"{e.__class__} from line 184")
            logging.error("Can't locate item_no/RefDesg/QtEach/PART_NO in Manex BOM!!")
            logging.error(f"{e}")
            print(e, "\n Can't locate item_no/RefDesg/QtEach/PART_NO in Manex BOM!!")
            sys.exit(1)
        manex_data = []
        error_msg = []
        for row in manex_rows[int(manex_dict["manex_start_row"])-1:int(manex_end_row)]:
            y1 = row[manex_col_des].value
            y2 = str(row[manex_col_qty].value)
            y3 = row[manex_col_partno].value
            y4 = row[manex_col_itemno].value
            if not bool(y1):
                error_msg.append(f"Missing designator for item no {y4}\n")
                logging.error(f"Missing designator for item no {y4}\n")
            if not bool(y2):
                error_msg.append(f"Missing quantity for item no {y4}\n")
                logging.error(f"Missing quantity for item no {y4}\n")
            elif not y2.isnumeric():
                error_msg.append(f"Non-integer quantity for item no {y4}\n")
                logging.error(f"Non-integer quantity for item no {y4}\n")
            elif int(y2) == 0:
                error_msg.append(f"Zero quantity for item no {y4}\n")
                logging.error(f"Zero quantity for item no {y4}\n")
            if not bool(y3):
                error_msg.append(f"Missing part-no for item no {y4}\n")
                logging.error(f"Missing part-no for item no {y4}\n")
            manex_data.append([y1, y2, y3, y4])
        wb_manex.close()
        if bool(error_msg):
            messagebox.showerror(title="Manex BOM Error", message="".join(error_msg))
            logging.error("Terminated!! Manex BOM not clean!!")
            sys.exit(1)
        return manex_data

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while reading Manex BOM!!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 184")
        logging.error("Error while reading Manex BOM!!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Manex BOM!!")
        sys.exit(1)


def Manex_data_interpretation(manex_dict, manex_data, separator_dict):
    try:
        for row in manex_data:
            y = row[0]
            if bool(y):
                if manex_dict["manex_delimiter"] is not None:
                    y = y.replace(" ", "").split(manex_dict["manex_delimiter"])
                    y = list(filter(None, y))
                rem = []
                new = []
                for item in y:
                    if manex_dict["manex_separator"] in item:
                        if item in separator_dict:
                            rem.append(item)
                            new.extend(separator_dict[item])
                        else:
                            res = []
                            stry = item.split(manex_dict["manex_separator"])
                            str1 = stry[0]
                            str2 = stry[1]
                            base = ""
                            for i in range(min(len(str1), len(str2))):
                                if str1[i] == str2[i]:
                                    base = f"{base}{str1[i]}"
                                else:
                                    break
                            if base == '':
                                break
                            count1 = 0
                            count2 = 0
                            for i in range(len(base) - 1):
                                if str1[i] == base[i]:
                                    count1 += 1
                                if str2[i] == base[i]:
                                    count2 += 1
                            str1 = str1[count1+1:]
                            str2 = str2[count2+1:]
                            my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                            my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                            if bool(my_list1) and bool(my_list2):
                                if len(my_list1) > 1:
                                    if my_list1[0].isalpha():
                                        for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                res.append(f"{base}{chr(i)}{j}")
                                    else:
                                        for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                            for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{i}{chr(j)}")
                                else:
                                    if my_list1[0].isalpha():
                                        for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            res.append(f"{base}{chr(i)}")
                                    else:
                                        for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                            res.append(f"{base}{j}")
                                rem.append(item)
                                new.extend(res)
                if bool(rem):
                    for obj in rem:
                        pointer = y.index(obj)
                        y.pop(pointer)
                    y.extend(new)
            row[0] = y
        logging.info("Manex data interpreted!!")
        return manex_data

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while interpreting Manex BOM!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 158")
        logging.error("Error while interpreting Manex BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while interpreting Manex BOM!")
        sys.exit(1)


def BOM_metadata():
    try:
        logging.info("Getting Customer BOM Info...")
        text = "BOM File Details"
        title = "Enter Details"
        input_list = ["Work Order No.", "RefDes Column", "Quantity Column", "Start Row", "End Row", "Delimiter", "Separator"]
        output = multenterbox(text, title, input_list)

        while output[1] is None or output[2] is None or output[3] is None or output[4] is None or not output[
            1].isalpha() or not output[2].isalpha() or not output[3].isnumeric() or not output[4].isnumeric():
            messagebox.showerror(title="Invalid Format", message="Please enter valid text formats")
            text = "BOM File Details"
            title = "Enter Details"
            input_list = ["Work Order No.", "RefDes Column", "Quantity Column", "Start Row", "End Row", "Delimiter", "Separator"]
            output = multenterbox(text, title, input_list)

        work_order = output[0].strip() if bool(output[0]) else 0
        bom_designator = (output[1].strip()).upper()
        bom_quantity = (output[2].strip()).upper()
        bom_start_row = int(output[3].strip())
        bom_end_row = int(output[4].strip())
        bom_delimiter = output[5].strip()
        bom_separator = output[6].strip()

        bom_dict = {"work_order": work_order, "bom_designator": bom_designator, "bom_quantity": bom_quantity,
                    "bom_start_row": bom_start_row, "bom_end_row": bom_end_row, "bom_delimiter": bom_delimiter, "bom_separator": bom_separator}
        logging.info("Info populated")
        return bom_dict

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while getting Customer BOM Detail Inputs!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 67")
        logging.error("Error while getting Customer BOM Detail Inputs!")
        logging.error(f"{e}")
        print(e, "\n Error while getting Customer BOM Detail Inputs!")
        sys.exit(1)


def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def BOM_data(Bom_File, bom_dict):

    messagebox.showinfo(title="Permission", message="Please close the BOM files if open. \nHit OK only when both files are closed!")
    try:
        file_extension = pathlib.Path(Bom_File).suffix
        logging.info("Reading Customer BOM Excel...")
        bom_data = []
        error_msg = []
        itemno = 1
        bom_col_des = ord(bom_dict["bom_designator"]) - 65
        bom_dict.update({"bom_col_des": bom_col_des})
        bom_col_qty = ord(bom_dict["bom_quantity"]) - 65
        if file_extension == ".xls" or file_extension == ".XLS":
            wb_bom = xlrd.open_workbook(Bom_File)
            ws_bom = wb_bom.sheet_by_index(0)
            for row in range(bom_dict["bom_start_row"] - 1, bom_dict["bom_end_row"]):
                var = ws_bom.row_values(row)
                x1 = var[bom_col_des]
                x2 = str(var[bom_col_qty])
                if not bool(x1):
                    error_msg.append(f"Missing designator for item no {itemno} or line no {row+1}\n")
                    logging.error(f"Missing designator for item no {itemno} or line no {row+1}\n")
                if not bool(x2):
                    error_msg.append(f"Missing quantity for item no {itemno} or line no {row+1}\n")
                    logging.error(f"Missing quantity for item no {itemno} or line no {row+1}\n")
                elif not x2.isnumeric():
                    if isfloat(x2):
                        if not float(x2).is_integer():
                            error_msg.append(f"Non-integer quantity for item no {itemno} or line no {row+1}\n")
                            logging.error(f"Non-integer quantity for item no {itemno} or line no {row+1}\n")
                        else:
                            x2 = int(float(x2))
                    else:
                        error_msg.append(f"Non-integer quantity for item no {itemno} or line no {row + 1}\n")
                        logging.error(f"Non-integer quantity for item no {itemno} or line no {row + 1}\n")
                bom_data.append([x1, x2, 0, row])
                itemno += 1

        else:
            wb_bom = openpyxl.load_workbook(Bom_File, data_only=True)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            count = bom_dict["bom_start_row"]
            for row in bom_rows[bom_dict["bom_start_row"] - 1:bom_dict["bom_end_row"]]:
                x1 = row[bom_col_des].value
                x2 = str(row[bom_col_qty].value)
                if not bool(x1):
                    error_msg.append(f"Missing designator for item no {itemno} or line no {count}\n")
                    logging.error(f"Missing designator for item no {itemno} or line no {count}\n")
                if not bool(x2):
                    error_msg.append(f"Missing quantity for item no {itemno} or line no {count}\n")
                    logging.error(f"Missing quantity for item no {itemno} or line no {count}\n")
                elif not x2.isnumeric():
                    if isfloat(x2):
                        if not float(x2).is_integer():
                            error_msg.append(f"Non-integer quantity for item no {itemno} or line no {count}\n")
                            logging.error(f"Non-integer quantity for item no {itemno} or line no {count}\n")
                        else:
                            x2 = int(float(x2))
                    else:
                        error_msg.append(f"Non-integer quantity for item no {itemno} or line no {count}\n")
                        logging.error(f"Non-integer quantity for item no {itemno} or line no {count}\n")
                bom_data.append([x1, x2, 0, count])
                itemno += 1
                count += 1
            wb_bom.close()
        if bool(error_msg):
            messagebox.showerror(title="Customer BOM Error", message="".join(error_msg))
            logging.error("Terminated!! Customer BOM not clean!!")
            sys.exit(1)
        return bom_data

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while reading Customer BOM File!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 122")
        logging.error("Error while reading Customer BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Customer BOM File!")
        sys.exit(1)


def Bom_data_interpretation(bom_data, bom_dict):
    try:
        separator_dict = {}
        separator_position = []
        for data in bom_data:
            res, sep_dict = readbom(data[0], bom_dict["bom_delimiter"], bom_dict["bom_separator"])
            if bool(sep_dict):
                separator_dict.update(sep_dict)
                separator_position.append([data[3], bom_data.index(data)])
            if len(res) != int(data[1]):
                qty = False
            else:
                qty = True
            data[0] = res
            data[2] = qty
            data.pop()
        return bom_data, separator_dict, separator_position

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while interpreting Customer BOM!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 122")
        logging.error("Error while interpreting Customer BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while interpreting Customer BOM File!")
        sys.exit(1)


def readbom(x, bom_delimiter, bom_separator):
    sep_dict = {}
    if bool(x):
        if bom_delimiter != '':
            x = x.replace(" ", "").split(bom_delimiter)
            x = list(filter(None, x))
        if bom_separator != '':
            rem = []
            new = []
            for item in x:
                if bom_separator in item:
                    is_sep = tk.messagebox.askyesno(title="Verify", message=f"Is {item} a separator?")
                    if is_sep is True:
                        res = []
                        text = f"{item}"
                        title = "Enter Details"
                        input_list = ["Base:", "From:", "To:"]
                        output = multenterbox(text, title, input_list)
                        base = output[0]
                        range_from = output[1]
                        range_to = output[2]
                        my_list1 = list(filter(None, re.split(r'(\d+)', range_from)))
                        my_list2 = list(filter(None, re.split(r'(\d+)', range_to)))
                        if base.isalpha():
                            base = base.upper()
                        elif not base.isnumeric():
                            base_list = list(base)
                            for val in base_list:
                                if val.isalpha():
                                    base_list[base_list.index(val)] = val.upper()
                            base = "".join(base_list)
                        if len(my_list1) > 1:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                        res.append(f"{base}{chr(i)}{j}")
                            else:
                                for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                    for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        res.append(f"{base}{i}{chr(j)}")
                        else:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    res.append(f"{base}{chr(i)}")
                            else:
                                for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                    res.append(f"{base}{j}")
                        sep_check = tk.messagebox.askyesno(title="Verify", message=f"{item} = {res}")
                        if sep_check is False:
                            res = []
                            text = f"{item}"
                            title = "Enter Details"
                            input_list = ["Base:", "From:", "To:"]
                            output = multenterbox(text, title, input_list)
                            base = output[0]
                            range_from = output[1]
                            range_to = output[2]
                            my_list1 = list(filter(None, re.split(r'(\d+)', range_from)))
                            my_list2 = list(filter(None, re.split(r'(\d+)', range_to)))
                            if base.isalpha():
                                base = base.upper()
                            elif not base.isnumeric():
                                base_list = list(base)
                                for val in base_list:
                                    if val.isalpha():
                                        base_list[base_list.index(val)] = val.upper()
                                base = "".join(base_list)
                            if len(my_list1) > 1:
                                if my_list1[0].isalpha():
                                    for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                            res.append(f"{base}{chr(i)}{j}")
                                else:
                                    for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                        for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            res.append(f"{base}{i}{chr(j)}")
                            else:
                                if my_list1[0].isalpha():
                                    for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        res.append(f"{base}{chr(i)}")
                                else:
                                    for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                        res.append(f"{base}{j}")
                            sep_check = tk.messagebox.askyesno(title="Verify", message=f"{item} = {res}")
                            if sep_check is False:
                                messagebox.showwarning(title="Attention", message=f"Please check manually for {item}!")
                                return x, sep_dict
                        rem.append(item)
                        new.extend(res)
                        sep_dict.update({item: res})
            for obj in rem:
                pointer = x.index(obj)
                x.pop(pointer)

            x.extend(new)
    return x, sep_dict


def Map_RefDes(bom_data, manex_data):
    try:
        logging.info("Mapping designators from Manex BOM to Customer BOM...")
        manex_pn = []
        duplicate = []
        # pcb = []
        for item in bom_data:
            if bool(item[0]):
                flag = 0
                pn = []
                for obj in manex_data:
                    if bool(obj[0]):
                        # if set("PCB").issubset(set(obj[0][0])):
                        #     if not bool(pcb):
                        #         pcb.append(obj[2])
                        if set(item[0]).issubset(set(obj[0])) or set(obj[0]).issubset(set(item[0])):
                            if obj[1] != 0:
                                if flag == 0:
                                    manex_pn.append(obj[2])
                                if obj[2] not in pn:
                                    pn.append(obj[2])
                                    flag += 1
                if flag == 0:
                    for obj in manex_data:
                        if bool(obj[0]):
                            if len(obj[0]) > 2:
                                rem = obj[0].pop()
                                if set(item[0]).issubset(set(obj[0])) or set(obj[0]).issubset(set(item[0])) or any(value in obj[0] for value in item[0]):
                                    if obj[1] != 0:
                                        if flag == 0:
                                            manex_pn.append(obj[2])
                                        if obj[2] not in pn:
                                            pn.append(obj[2])
                                            flag += 1
                                obj[0].append(rem)
                # if flag == 0:
                #     if set('PCB').issubset(set(item[0][0])):
                #         for obj in manex_data:
                #             if bool(obj[0]) and len(obj[0]) == 1:
                #                 if set('PCB').issubset(set(obj[0][0])):
                #                     if obj[1] != 0:
                #                         if flag == 0:
                #                             manex_pn.append(obj[2])
                #                         if obj[2] not in pn:
                #                             pn.append(obj[2])
                #                             flag += 1
                if flag == 0:
                    manex_pn.append('Not in Manex')
                if flag > 1:
                    duplicate.extend(pn)
            else:
                manex_pn.append(None)
        logging.info("Finished mapping")
        return manex_pn, duplicate

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while mapping designators from Manex BOM to Customer BOM!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 185")
        logging.error("Error while mapping designators from Manex BOM to Customer BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from Manex BOM to Customer BOM!")
        sys.exit(1)


def Write_Bom(Bom_File, bom_data, bom_dict, manex_pn, duplicate, separator_position):
    try:
        logging.info("Writing to Customer Bom Excel...")
        # file_extension = pathlib.Path(Bom_File).suffix
        # if file_extension == ".xls" or file_extension == ".XLS":
        xlApp = client.Dispatch("Excel.Application")
        wkbk = xlApp.Workbooks.open(Bom_File)
        wksht = wkbk.Worksheets(1)
        wksht.Columns("A").EntireColumn.Insert()

        j = 0
        for i in range(bom_dict["bom_start_row"], bom_dict["bom_end_row"]+1):
            wksht.Cells(i, 1).Value = manex_pn[j]
            j += 1
        wksht.Columns("A").EntireColumn.Insert()
        wksht.Rows(bom_dict["bom_end_row"] + 1).EntireRow.Insert()
        wksht.Rows(bom_dict["bom_end_row"] + 1).EntireRow.Insert()
        wksht.Cells(bom_dict["bom_end_row"] + 1, 3).Interior.ColorIndex = 0
        wksht.Cells(bom_dict["bom_end_row"] + 2, 3).Interior.ColorIndex = 0
        # wksht.Cells(bom_end_row + 2, 2).Value = "Last modified at"
        string = f"Manex PN added on {datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S')}"
        wksht.Cells(bom_dict["bom_end_row"] + 2, 2).Value = string
        # if bool(pcb) and pcb[0] not in manex_pn:
        #     wksht.Rows(bom_dict["bom_end_row"]+1).EntireRow.Insert()
        #     wksht.Cells(bom_dict["bom_end_row"] + 1, 3).Interior.ColorIndex = 0
        #     wksht.Cells(bom_dict["bom_end_row"] + 1, bom_dict["bom_col_des"]+3).Value = "PCB"
        #     wksht.Cells(bom_dict["bom_end_row"] + 1, 2).Value = pcb[0]
        pointer = 0
        for i in range(bom_dict["bom_start_row"], bom_dict["bom_end_row"]+1):
            if wksht.Cells(i, 2).Value == "Not in Manex":
                wksht.Cells(i, 1).Value = "Check"
                wksht.Cells(i, 1).Interior.ColorIndex = 6
            elif wksht.Cells(i, 2).Value in duplicate:
                wksht.Cells(i, 1).Value = "Duplicate RefDesgs in BOM"
                wksht.Cells(i, 1).Interior.ColorIndex = 8
            if not bool(bom_data[pointer][2]):
                wksht.Cells(i, 1).Value = "Quantity Column and RefDesg Count does not match"
            if not bool(bom_data[pointer][1]):
                for col in range(1, int(wksht.UsedRange.Columns.Count)):
                    wksht.Cells(i, col).Font.ColorIndex = 3
            pointer += 1

        file_extension = pathlib.Path(Bom_File).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            for value in separator_position:
                wksht.Cells(value[0]+1, bom_dict["bom_col_des"]+3).Value = ", ".join(bom_data[value[1]][0])
        else:
            for value in separator_position:
                wksht.Cells(value[0], bom_dict["bom_col_des"]+3).Value = ", ".join(bom_data[value[1]][0])

        wkbk.Save()
        wkbk.Close(True)
        xlApp.Quit()

        if bool(bom_dict["work_order"]):
            name = Bom_File.split("/")
            filename = name.pop()
            filename = f"{bom_dict['work_order']}_{filename}"
            name.append(filename)
            new_name = "/".join(name)
            os.rename(Bom_File, new_name)

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while writing Customer BOM!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 242")
        logging.error("Error while writing Customer BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while writing Customer BOM!")
        # if file_extension == ".xls" or file_extension == ".XLS":
        if bool(wkbk):
            wkbk.Close()
            xlApp.Quit()
        # else:
        #     if bool(wb_bom):
        #         wb_bom.close()
        sys.exit(1)


def Write_Manex(Manex_File, manex_dict, manex_pn, duplicate):
    try:
        logging.info("Writing to Manex BOM Excel...")
        wb_manex = openpyxl.load_workbook(Manex_File, data_only=True)
        ws_manex = wb_manex.worksheets[0]
        r = manex_dict["manex_start_row"]
        for rows in ws_manex.iter_rows(min_row=manex_dict["manex_start_row"], max_row=manex_dict["manex_end_row"], min_col=1, max_col=20):
            if ws_manex[f"{chr(manex_dict['manex_col_partno']+65)}{str(r)}"].value not in manex_pn:
                for cell in rows:
                    cell.font = Font(color="00FF1414")
            if ws_manex[f"{chr(manex_dict['manex_col_partno']+65)}{str(r)}"].value in duplicate:
                for cell in rows:
                    cell.font = Font(color="000096FF")
            r += 1
        logging.info("Finished writing")
        wb_manex.save(Manex_File)

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message=f"Error while writing Manex BOM!\n\n{e.__class__}\n{e}")
        logging.error(f"{e.__class__} from line 242")
        logging.error("Error while writing Manex BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while writing Manex BOM!")
        if bool(wb_manex):
            wb_manex.close()
        sys.exit(1)


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")
    Manex_File, Bom_File = getfiles()

    manex_dict = manex_metadata()
    manex_data = Manex_data(Manex_File, manex_dict)
    bom_dict = BOM_metadata()
    bom_data = BOM_data(Bom_File, bom_dict)

    bom_data_manipulated, separator_dict, separator_position = Bom_data_interpretation(bom_data, bom_dict)
    manex_data_manipulated = Manex_data_interpretation(manex_dict, manex_data, separator_dict)

    manex_pn, duplicate = Map_RefDes(bom_data_manipulated, manex_data_manipulated)

    Write_Bom(Bom_File, bom_data_manipulated, bom_dict, manex_pn, duplicate, separator_position)
    Write_Manex(Manex_File, manex_dict, manex_pn, duplicate)

    logging.info("Successfully Executed!!!\n\n")
    print("Successfully Executed!!!")
    messagebox.showinfo(title="Status", message="Completed!!!")
