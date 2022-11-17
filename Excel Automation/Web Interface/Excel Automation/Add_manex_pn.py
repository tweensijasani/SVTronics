import re
import xlrd
import pathlib
import logging
import openpyxl
import datetime
import configparser
import tkinter as tk
from easygui import *
from win32com import client
from tkinter import messagebox
from openpyxl.styles import PatternFill, Font

logging.basicConfig(level=logging.DEBUG, filename="Excel_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")


def MapDes(customer_bom, manex_bom, bom_data, file_extension, start_row, end_row, wb_bom, ws_bom, bom_col_des):

    manex_data, manex_start_row, manex_end_row, ws_manex, wb_manex, manex_partno = ManexInfo(manex_bom)

    try:
        logging.info("Mapping designators from Manex BOM to Customer BOM...")
        manex_pn = []
        duplicate = []
        pcb = []
        for item in bom_data:
            if item[0] is not None:
                flag = 0
                pn = []
                for obj in manex_data:
                    if bool(obj[0]):
                        if set("PCB").issubset(set(obj[0][0])):
                            if not bool(pcb):
                                pcb.append(obj[2])
                        if set(item[0]).issubset(set(obj[0])) or set(obj[0]).issubset(set(item[0])):
                            if obj[1] != 0:
                                if flag == 0:
                                    manex_pn.append(obj[2])
                                if obj[2] not in pn:
                                    pn.append(obj[2])
                                    flag += 1
                if flag == 0:
                    for obj in manex_data:
                        if obj[0] is not None:
                            if len(obj[0]) > 2:
                                rem = obj[0].pop()
                                if set(item[0]).issubset(set(obj[0])) or set(obj[0]).issubset(set(item[0])) or any(value in obj[0] for value in item[0]):
                                    if obj[1] != 0:
                                        if flag == 0:
                                            manex_pn.append(obj[2])
                                        if obj[2] not in pn:
                                            pn.append(obj[2])
                                            flag += 1
                                obj[0].append(rem)
                if flag == 0:
                    if set('PCB').issubset(set(item[0][0])):
                        for obj in manex_data:
                            if obj[0] is not None and len(obj[0]) == 1:
                                if set('PCB').issubset(set(obj[0][0])):
                                    if obj[1] != 0:
                                        if flag == 0:
                                            manex_pn.append(obj[2])
                                        if obj[2] not in pn:
                                            pn.append(obj[2])
                                            flag += 1
                if flag == 0:
                    manex_pn.append('Manex PN not found')
                if flag > 1:
                    duplicate.extend(pn)
            else:
                manex_pn.append(None)
        logging.info("Finished mapping")
        cust = WriteCustBom(customer_bom, file_extension, start_row, end_row, wb_bom, ws_bom, manex_pn, duplicate, pcb, bom_col_des)
        man = WriteManexBom(manex_bom, manex_start_row, manex_end_row, ws_manex, wb_manex, manex_partno, manex_pn, duplicate)
        return cust and man

    except Exception as e:
        logging.error(f"{e.__class__} from line 22")
        logging.error("Error while mapping designators from Manex BOM to Customer BOM!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from Manex BOM to Customer BOM!")
        return False


def ReadCustBom(customer_bom, manex_bom, designator, quantity, start_row, end_row, delimiter, separator):

    try:
        file_extension = pathlib.Path(customer_bom).suffix
        logging.info("Reading Customer BOM Excel...")
        if file_extension == ".xls":
            wb_bom = xlrd.open_workbook(customer_bom)
            ws_bom = wb_bom.sheet_by_index(0)
            bom_data = []
            bom_col_des = ord(designator) - 65
            bom_col_qty = ord(quantity) - 65
            for row in range(start_row-1, end_row):
                var = ws_bom.row_values(row)
                x = var[bom_col_des]
                res = CustBomInfo(x, delimiter, separator)
                bom_data.append([res, var[bom_col_qty]])
        else:
            wb_bom = openpyxl.load_workbook(customer_bom)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            bom_data = []
            bom_col_des = ord(designator)-65
            bom_col_qty = ord(quantity)-65
            for row in bom_rows[int(start_row)-1:int(end_row)]:
                x = row[bom_col_des].value
                res = CustBomInfo(x, delimiter, separator)
                bom_data.append([res, row[bom_col_qty].value])
        logging.info("Finished reading")
        return MapDes(customer_bom, manex_bom, bom_data, file_extension, start_row, end_row, wb_bom, ws_bom, bom_col_des)

    except Exception as e:
        logging.error(f"{e.__class__} from line 85")
        logging.error("Error while reading Customer BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Customer BOM File!")
        return False


def CustBomInfo(x, bom_delimiter, bom_separator):
    root = tk.Tk()
    root.withdraw()
    root.lift()
    root.attributes('-topmost', True)
    if x is not None:
        if bom_delimiter is not None:
            x = x.replace(" ", "").split(bom_delimiter)
            x = list(filter(None, x))
        if bom_separator is not None:
            for item in x:
                if bom_separator in item:
                    is_sep = tk.messagebox.askyesno(title="Verify", message=f"Is {item} a separator?")
                    if is_sep is True:
                        res = []
                        text = f"{item}"
                        title = "Enter Details"
                        input_list = ["Base:", "From:", "To:"]
                        output = multenterbox(text, title, input_list)
                        base = output[0].strip()
                        range_from = output[1].strip()
                        range_to = output[2].strip()
                        my_list1 = list(filter(None, re.split(r'(\d+)', range_from)))
                        my_list2 = list(filter(None, re.split(r'(\d+)', range_to)))
                        if base.isalpha():
                            base = base.upper()
                        if len(my_list1) > 1:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                        res.append(f"{base}{chr(i)}{j}")
                            else:
                                for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                    for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        res.append(f"{base}{i}{chr(j)}")
                        else:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    res.append(f"{base}{chr(i)}")
                            else:
                                for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                    res.append(f"{base}{j}")
                        sep_check = tk.messagebox.askyesno(title="Verify", message=f"{item} = {res}")
                        if sep_check is False:
                            res = []
                            text = f"{item}"
                            title = "Enter Details"
                            input_list = ["Base:", "From:", "To:"]
                            output = multenterbox(text, title, input_list)
                            base = output[0].strip()
                            range_from = output[1].strip()
                            range_to = output[2].strip()
                            my_list1 = list(filter(None, re.split(r'(\d+)', range_from)))
                            my_list2 = list(filter(None, re.split(r'(\d+)', range_to)))
                            if base.isalpha():
                                base = base.upper()
                            elif not base.isnumeric():
                                base_list = list(filter(None, re.split(r'(\d+)', base)))
                                if base_list[0].isalpha():
                                    base_list[0] = base_list[0].upper()
                                else:
                                    base_list[1] = base_list[1].upper()
                                base = f"{base_list[0]}{base_list[1]}"
                            if len(my_list1) > 1:
                                if my_list1[0].isalpha():
                                    for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                            res.append(f"{base}{chr(i)}{j}")
                                else:
                                    for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                        for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            res.append(f"{base}{i}{chr(j)}")
                            else:
                                if my_list1[0].isalpha():
                                    for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        res.append(f"{base}{chr(i)}")
                                else:
                                    for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                        res.append(f"{base}{j}")
                            sep_check = tk.messagebox.askyesno(title="Verify", message=f"{item} = {res}")
                            if sep_check is False:
                                messagebox.showwarning(title="Attention", message=f"Please check manually for {item}!")
                                return x
                        pointer = x.index(item)
                        x.pop(pointer)
                        x.extend(res)
    root.destroy()
    return x


def ManexInfo(manex_bom):

    try:
        logging.info("Reading setup.ini file...")
        config = configparser.ConfigParser()
        config.read('setup.ini')

        manex_designator = config['INITIALIZATION']['RefDes']
        manex_quantity = config['INITIALIZATION']['Quantity']
        manex_partno = config['INITIALIZATION']['PartNumber']
        manex_start_row = int(config['INITIALIZATION']['StartRow'])
        manex_delimiter = config['INITIALIZATION']['Delimiter']
        manex_separator = config['INITIALIZATION']['Separator']
        logging.info("Info populated")

    except Exception as e:
        logging.error(f"{e.__class__} from line 168")
        logging.error("Error while getting Manex BOM Detail Inputs!")
        logging.error(f"{e}")
        print(e, "\n Error while getting Manex BOM Detail Inputs!")
        return False

    try:
        logging.info("Reading Manex BOM Excel...")
        wb_manex = openpyxl.load_workbook(manex_bom)
        ws_manex = wb_manex.worksheets[0]
        manex_end_row = int(ws_manex.max_row)
        manex_rows = list(ws_manex.rows)
        manex_data = []
        manex_col_des = ord(manex_designator) - 65
        manex_col_qty = ord(manex_quantity) - 65
        manex_col_partno = ord(manex_partno) - 65
        for row in manex_rows[int(manex_start_row)-1:int(manex_end_row)]:
            y = row[manex_col_des].value
            if y is not None:
                if manex_delimiter is not None:
                    y = (row[manex_col_des].value).replace(" ", "").split(manex_delimiter)
                    y = list(filter(None, y))
                for item in y:
                    if manex_separator in item:
                        res = []
                        stry = item.split(manex_separator)
                        str1 = stry[0]
                        str2 = stry[1]
                        base = ""
                        for i in range(len(str1) - 1):
                            if str1[i] == str2[i]:
                                base = f"{base}{str1[i]}"
                            else:
                                break
                        str1 = str1.lstrip(base)
                        str2 = str2.lstrip(base)
                        my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                        my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                        if len(my_list1) > 1:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                        res.append(f"{base}{chr(i)}{j}")
                            else:
                                for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                    for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                        res.append(f"{base}{i}{chr(j)}")
                        else:
                            if my_list1[0].isalpha():
                                for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                    res.append(f"{base}{chr(i)}")
                            else:
                                for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                    res.append(f"{base}{j}")
                        pointer = y.index(item)
                        y.pop(pointer)
                        y.extend(res)
            manex_data.append([y, row[manex_col_qty].value, row[manex_col_partno].value])
        logging.info("Finished reading")
        return manex_data, manex_start_row, manex_end_row, ws_manex, wb_manex, manex_partno

    except Exception as e:
        logging.error(f"{e.__class__} from line 190")
        logging.error("Error while reading Manex BOM File!")
        logging.error(f"{e}")
        print(e, "\n Error while reading Manex BOM File!")
        return False


def WriteCustBom(customer_bom, file_extension, bom_start_row, bom_end_row, wb_bom, ws_bom, manex_pn, duplicate, pcb, bom_col_des):

    try:
        logging.info("Writing to Customer Bom Excel...")
        if file_extension == ".xls":
            xlApp = client.Dispatch("Excel.Application")
            wkbk = xlApp.Workbooks.open(customer_bom)
            wksht = wkbk.Worksheets(1)
            wksht.Columns("A").EntireColumn.Insert()
            j = 0
            for i in range(bom_start_row, bom_end_row+1):
                wksht.Cells(i, 1).Value = manex_pn[j]
                j += 1
            wksht.Columns("A").EntireColumn.Insert()
            for i in range(bom_start_row, bom_end_row+1):
                if wksht.Cells(i, 2).Value == "Manex PN not found":
                    wksht.Cells(i, 1).Value = "Check"
                    wksht.Cells(i, 1).Interior.ColorIndex = 6
                elif wksht.Cells(i, 2).Value in duplicate:
                    wksht.Cells(i, 1).Value = "Duplicate"
                    wksht.Cells(i, 1).Interior.ColorIndex = 8
                elif wksht.Cells(i, 2).Value is None:
                    wksht.Cells(i, 1).Value = "Ref des Missing"
                    wksht.Cells(i, 1).Interior.ColorIndex = 6
            wksht.Rows(bom_end_row + 1).EntireRow.Insert()
            wksht.Rows(bom_end_row + 1).EntireRow.Insert()
            wksht.Cells(bom_end_row + 1, 1).Interior.ColorIndex = 0
            wksht.Cells(bom_end_row + 1, 3).Interior.ColorIndex = 0
            wksht.Cells(bom_end_row + 2, 1).Interior.ColorIndex = 0
            wksht.Cells(bom_end_row + 2, 3).Interior.ColorIndex = 0
            wksht.Cells(bom_end_row + 2, 1).Value = "Last modified at"
            wksht.Cells(bom_end_row + 2, 3).Value = str(datetime.date.today())
            wksht.Cells(bom_end_row + 2, 4).Value = str(datetime.datetime.now().strftime("%H:%M:%S"))
            if bool(pcb) and pcb[0] not in manex_pn:
                wksht.Rows(bom_end_row + 1).EntireRow.Insert()
                wksht.Cells(bom_end_row + 1, 1).Interior.ColorIndex = 0
                wksht.Cells(bom_end_row + 1, 3).Interior.ColorIndex = 0
                wksht.Cells(bom_end_row + 1, bom_col_des + 3).Value = "PCB"
                wksht.Cells(bom_end_row + 1, 2).Value = pcb[0]
            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()
        else:
            ws_bom.insert_cols(0)
            i = 0
            for row_num in range(int(bom_start_row), int(bom_end_row)+1):
                ws_bom.cell(row=row_num, column=1).value = manex_pn[i]
                i += 1

            ws_bom.insert_cols(0)
            r = bom_start_row
            for rows in ws_bom.iter_rows(min_row=bom_start_row, max_row=bom_end_row, min_col=1, max_col=20):
                if ws_bom[f"B{str(r)}"].value == "Manex PN not found":
                    rows[0].fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
                    rows[0].value = "Check"
                elif ws_bom[f"B{str(r)}"].value in duplicate:
                    rows[0].fill = PatternFill(start_color="000096FF", end_color="000096FF", fill_type="solid")
                    rows[0].value = "Duplicate"
                elif ws_bom[f"B{str(r)}"].value is None:
                    rows[0].fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
                    rows[0].value = "Ref des Missing"
                r += 1
            ws_bom.insert_rows(bom_end_row + 1)
            ws_bom.insert_rows(bom_end_row + 1)
            ws_bom.cell(row=bom_end_row + 2, column=1).value = "Last modified at"
            ws_bom.cell(row=bom_end_row + 2, column=3).value = str(datetime.datetime.now())
            if bool(pcb) and pcb[0] not in manex_pn:
                ws_bom.insert_rows(bom_end_row + 1)
                ws_bom.cell(row=bom_end_row + 1, column=bom_col_des + 3).value = "PCB"
                ws_bom.cell(row=bom_end_row + 1, column=2).value = pcb[0]
            wb_bom.save(customer_bom)
        logging.info("Finished writing")
        return True

    except Exception as e:
        logging.error(f"{e.__class__} from line 220")
        logging.error("Error while writing Customer BOM excel!")
        logging.error(f"{e}")
        print(e, "\n Error while writing Customer BOM excel!")
        return False


def WriteManexBom(manex_bom, manex_start_row, manex_end_row, ws_manex, wb_manex, manex_partno, manex_pn, duplicate):

    try:
        logging.info("Writing to Manex BOM Excel...")
        r = manex_start_row
        for rows in ws_manex.iter_rows(min_row=manex_start_row, max_row=manex_end_row, min_col=1, max_col=20):
            if ws_manex[f"{manex_partno}{str(r)}"].value not in manex_pn:
                for cell in rows:
                    cell.font = Font(color="00FF1414")
            if ws_manex[f"{manex_partno}{str(r)}"].value in duplicate:
                for cell in rows:
                    cell.font = Font(color="000096FF")
            r += 1
        logging.info("Finished writing")
        wb_manex.save(manex_bom)
        return True

    except Exception as e:
        logging.error(f"{e.__class__} from line 279")
        logging.error("Error while writing Manex file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing Manex file!")
        return False


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")
    logging.info("Sucessfully Executed!!!\n\n")
    print("Sucessfully Executed!!!")

