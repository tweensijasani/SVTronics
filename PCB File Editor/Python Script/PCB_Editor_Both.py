import re
import sys
import xlrd
import pathlib
import logging
import openpyxl
import easygui.boxes
import tkinter as tk
from easygui import *
from win32com import client
from tkinter import filedialog, messagebox

logging.basicConfig(level=logging.DEBUG, filename="PCB_both_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")

root = tk.Tk()
root.withdraw()

easygui.boxes.global_state.PROPORTIONAL_FONT_FAMILY = ("MS", "Arial")
easygui.boxes.global_state.MONOSPACE_FONT_FAMILY = "Arial"
easygui.boxes.global_state.PROPORTIONAL_FONT_SIZE = 14
easygui.boxes.global_state.MONOSPACE_FONT_SIZE = 14
easygui.boxes.global_state.TEXT_ENTRY_FONT_SIZE = 14
easygui.boxes.global_state.prop_font_line_length = 30
easygui.boxes.global_state.fixw_font_line_length = 40


def getfiles():
    try:
        counter = 0
        bot_file = filedialog.askopenfilename(title="Select Bottom PCB File")
        while bot_file == '' and counter < 1:
            messagebox.showerror(title="File Error", message="PCB BOT File Not Selected!")
            logging.warning("PCB BOT File Not Selected!")
            bot_file = filedialog.askopenfilename(title="Select Bottom PCB File")
            counter += 1
        if bot_file == '':
            messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
            logging.error("Terminated: PCB Bottom not selected!")
            print("Terminated: PCB Bottom not selected!")
            sys.exit(1)
        else:
            logging.info("PCB Bottom File Selected!")

        counter = 0
        top_file = filedialog.askopenfilename(title="Select Top PCB File")
        while top_file == '' and counter < 1:
            messagebox.showerror(title="File Error", message="PCB TOP File Not Selected!")
            logging.warning("PCB TOP File Not Selected!")
            top_file = filedialog.askopenfilename(title="Select Top PCB File")
            counter += 1
        if top_file == '':
            messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
            logging.error("Terminated: PCB Top not selected!")
            print("Terminated: PCB Top not selected!")
            sys.exit(1)
        else:
            logging.info("PCB Top File Selected!")

        counter = 0
        Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                              filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
        while Bom_File == '' and counter < 1:
            messagebox.showerror(title="File Error", message="BOM File Not Selected!")
            logging.warning("BOM File Not Selected!")
            Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                                  filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
            counter += 1
        if Bom_File == '':
            messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
            logging.error("Terminated: Customer BOM not selected!")
            print("Terminated: Customer BOM not selected!")
            sys.exit(1)
        else:
            logging.info("Customer BOM Excel Selected!")
        return Bom_File, bot_file, top_file

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__}")
        logging.error("Error while fetching files!")
        logging.error(f"{e}")
        print(e, "\n Error while fetching files!")
        sys.exit(1)


def read_pcb(bot_file, top_file):
    try:
        logging.info("Reading bot.pcb file...")
        botrefdes = []
        bot_textfile = open(bot_file, 'r')
        botfiledata = bot_textfile.readlines()
        for line in botfiledata:
            match = re.match("^F9\s", line)
            if match:
                botrefdes.append([line.strip().split(" ").pop()])

        logging.info("Reading top.pcb file...")
        toprefdes = []
        top_textfile = open(top_file, 'r')
        topfiledata = top_textfile.readlines()
        for line in topfiledata:
            match = re.match("^F9\s", line)
            if match:
                toprefdes.append([line.strip().split(" ").pop()])
        logging.info("Finished reading pcb files!")
        bot_textfile.close()
        top_textfile.close()
        return botrefdes, toprefdes

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 80")
        logging.error("Error while reading .pcb files!")
        logging.error(f"{e}")
        print(e, "\n Error while reading .pcb files!")
        sys.exit(1)


def BOM_metadata():
    try:
        logging.info("Getting BOM column info...")
        text = "BOM File Details"
        title = "Enter Details"
        input_list = ["RefDes Column", "Manex P/N Column", "Start Row", "End Row", "Delimiter", "Separator"]
        output = multenterbox(text, title, input_list)

        while output[0] is None or output[1] is None or output[2] is None or output[3] is None or not output[0].isalpha() \
                or not output[1].isalpha() or not output[2].isnumeric() or not output[3].isnumeric():
            messagebox.showerror(title="Invalid Format", message="Please enter valid text formats")
            text = "BOM File Details"
            title = "Enter Details"
            input_list = ["RefDes Column", "Manex P/N Column", "Start Row", "End Row", "Delimiter", "Separator"]
            output = multenterbox(text, title, input_list)

        bom_designator = (output[0].strip()).upper()
        bom_pn = (output[1].strip()).upper()
        bom_start_row = int(output[2].strip())
        bom_end_row = int(output[3].strip())
        bom_delimiter = output[4].strip()
        bom_separator = output[5].strip()

        bom_dict = {"bom_designator": bom_designator, "bom_pn": bom_pn, "bom_start_row": bom_start_row, "bom_end_row": bom_end_row,
                    "bom_delimiter": bom_delimiter, "bom_separator": bom_separator}
        logging.info("Info populated!")
        return bom_dict

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 130")
        logging.error("Error while getting BOM info!")
        logging.error(f"{e}")
        print(e, "\n Error while getting BOM info!")
        sys.exit(1)


def readBom(bomfile, bom_dict):
    try:
        messagebox.showinfo(title="Permission",
                            message="Please close the BOM file if open. \nHit OK only when the file is closed!")

        logging.info("Reading BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            wb_bom = xlrd.open_workbook(bomfile)
            ws_bom = wb_bom.sheet_by_index(0)
            bot_bom_data = []
            top_bom_data = []
            bom_col_des = ord(bom_dict["bom_designator"]) - 65
            bom_col_pn = ord(bom_dict["bom_pn"]) - 65
            for row in range(bom_dict["bom_start_row"]-1, bom_dict["bom_end_row"]):
                var = ws_bom.row_values(row)
                x = var[bom_col_des]
                if bool(x):
                    if bom_dict["bom_delimiter"] != '':
                        x = (var[bom_col_des]).replace(" ", "").split(bom_dict["bom_delimiter"])
                        x = list(filter(None, x))
                        if bom_dict["bom_separator"] != '':
                            for item in x:
                                if bom_dict["bom_separator"] in item:
                                    res = []
                                    stry = item.split(bom_dict["bom_separator"])
                                    str1 = stry[0]
                                    str2 = stry[1]
                                    base = ""
                                    for i in range(len(str1) - 1):
                                        if str1[i] == str2[i]:
                                            base = f"{base}{str1[i]}"
                                        else:
                                            break
                                    if base == '':
                                        break
                                    count1 = 0
                                    count2 = 0
                                    for i in range(len(base) - 1):
                                        if str1[i] == base[i]:
                                            count1 += 1
                                        if str2[i] == base[i]:
                                            count2 += 1
                                    str1 = str1[count1 + 1:]
                                    str2 = str2[count2 + 1:]
                                    my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                                    my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                                    if len(my_list1) > 1:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                    res.append(f"{base}{chr(i)}{j}")
                                        else:
                                            for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                    res.append(f"{base}{i}{chr(j)}")
                                    else:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{chr(i)}")
                                        else:
                                            for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                                res.append(f"{base}{j}")
                                    pointer = x.index(item)
                                    x.pop(pointer)
                                    x.extend(res)
                y = str(var[bom_col_pn])
                if bool(y):
                    y = y.split(".")
                    bot_bom_data.append([x, y[0], 0])
                    top_bom_data.append([x, y[0], 0])
                else:
                    bot_bom_data.append([x, y, 0])
                    top_bom_data.append([x, y, 0])
        else:
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            bot_bom_data = []
            top_bom_data = []
            bom_col_des = ord(bom_dict["bom_designator"]) - 65
            bom_col_pn = ord(bom_dict["bom_pn"]) - 65
            for row in bom_rows[int(bom_dict["bom_start_row"]) - 1:int(bom_dict["bom_end_row"])]:
                x = row[bom_col_des].value
                if bool(x):
                    if bom_dict["bom_delimiter"] != '':
                        x = (row[bom_col_des].value).replace(" ", "").split(bom_dict["bom_delimiter"])
                        x = list(filter(None, x))
                        if bom_dict["bom_separator"] != '':
                            for item in x:
                                if bom_dict["bom_separator"] in item:
                                    res = []
                                    stry = item.split(bom_dict["bom_separator"])
                                    str1 = stry[0]
                                    str2 = stry[1]
                                    base = ""
                                    for i in range(len(str1) - 1):
                                        if str1[i] == str2[i]:
                                            base = f"{base}{str1[i]}"
                                        else:
                                            break
                                    if base == '':
                                        break
                                    count1 = 0
                                    count2 = 0
                                    for i in range(len(base) - 1):
                                        if str1[i] == base[i]:
                                            count1 += 1
                                        if str2[i] == base[i]:
                                            count2 += 1
                                    str1 = str1[count1 + 1:]
                                    str2 = str2[count2 + 1:]
                                    my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                                    my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                                    if len(my_list1) > 1:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                    res.append(f"{base}{chr(i)}{j}")
                                        else:
                                            for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                    res.append(f"{base}{i}{chr(j)}")
                                    else:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{chr(i)}")
                                        else:
                                            for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                                res.append(f"{base}{j}")
                                    pointer = x.index(item)
                                    x.pop(pointer)
                                    x.extend(res)
                y = str(row[bom_col_pn].value)
                if bool(y):
                    y = y.split(".")
                    bot_bom_data.append([x, y[0], 0])
                    top_bom_data.append([x, y[0], 0])
                else:
                    bot_bom_data.append([x, y, 0])
                    top_bom_data.append([x, y, 0])
            wb_bom.close()
        logging.info("Finished reading BOM excel!")
        return bot_bom_data, top_bom_data

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 130")
        logging.error("Error while reading BOM Excel!")
        logging.error(f"{e}")
        print(e, "\n Error while reading BOM Excel!")
        sys.exit(1)


def mapping(botrefdes, toprefdes, bot_bom_data, top_bom_data):
    try:
        logging.info("Mapping bot.pcb RefDes to BOM excel...")
        pointer = 0
        for value in botrefdes:
            flag = 0
            for ref in bot_bom_data:
                if bool(ref[0]) and value[0] in ref[0] and ref[1] is not None:
                    botrefdes[pointer].append(ref[1])
                    ref[2] = 1
                    flag = 1
                    break
            if flag == 0:
                botrefdes[pointer].append("Not found")
            pointer += 1

        logging.info("Mapping top.pcb RefDes to BOM excel...")
        pointer = 0
        for value in toprefdes:
            flag = 0
            for ref in top_bom_data:
                if bool(ref[0]) and value[0] in ref[0] and ref[1] is not None:
                    toprefdes[pointer].append(ref[1])
                    ref[2] = 1
                    flag = 1
                    break
            if flag == 0:
                toprefdes[pointer].append("Not found")
            pointer += 1
        logging.info("Finished mapping!")
        return botrefdes, toprefdes, bot_bom_data, top_bom_data

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 192")
        logging.error("Error while mapping designators from .pcb files to excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from .pcb files to excel file!")
        sys.exit(1)


def Write_PCB(bot_file, top_file, botrefdes, toprefdes):
    try:
        logging.info("Editing bot.pcb content...")
        pointer = 0
        line_pointer = 0
        bot_textfile = open(bot_file, 'r')
        botfiledata = bot_textfile.readlines()
        for line in botfiledata:
            match = re.match("^F8\s", line)
            if match:
                if botrefdes[pointer][1] != "Not found" and botrefdes[pointer][1] is not None:
                    x = line.strip().split(" ")
                    x = x[0:7]
                    x.append(botrefdes[pointer][1])
                    x = " ".join(x)
                    botfiledata[line_pointer] = x + "\n"
                pointer += 1
            line_pointer += 1

        logging.info("Editing top.pcb content...")
        pointer = 0
        line_pointer = 0
        top_textfile = open(top_file, 'r')
        topfiledata = top_textfile.readlines()
        for line in topfiledata:
            match = re.match("^F8\s", line)
            if match:
                if toprefdes[pointer][1] != "Not found" and toprefdes[pointer][1] is not None:
                    x = line.strip().split(" ")
                    x = x[0:7]
                    x.append(toprefdes[pointer][1])
                    x = " ".join(x)
                    topfiledata[line_pointer] = x + "\n"
                pointer += 1
            line_pointer += 1
        logging.info("Finished editing!")

        bot_textfile.close()
        top_textfile.close()

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 231")
        logging.error("Error while editing .pcb file content!")
        logging.error(f"{e}")
        print(e, "\n Error while editing .pcb file content!")
        sys.exit(1)

    try:
        logging.info("Writing modified bot.pcb file...")
        var = bot_textfile.name.split("/")
        temp = var.pop()
        new_file_name = temp.replace(".pcb", "_modified.pcb")
        new_file_name = new_file_name.replace(".PCB", "_modified.PCB")
        var.append(new_file_name)
        new_file = "/".join(var)

        with open(new_file, "w") as f:
            for item in botfiledata:
                f.write("%s" % item)
        f.close()

        logging.info("Writing modified top.pcb file...")
        var = top_textfile.name.split("/")
        temp = var.pop()
        new_file_name = temp.replace(".pcb", "_modified.pcb")
        new_file_name = new_file_name.replace(".PCB", "_modified.PCB")
        var.append(new_file_name)
        new_file = "/".join(var)

        with open(new_file, "w") as f:
            for item in topfiledata:
                f.write("%s" % item)
        f.close()
        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 272")
        logging.error("Error while creating modified .pcb files!")
        logging.error(f"{e}")
        print(e, "\n Error while creating modified .pcb files!")
        sys.exit(1)


def Write_Bom(bomfile, bot_bom_data, top_bom_data, botrefdes, toprefdes):
    try:
        logging.info("Writing to BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            xlApp = client.Dispatch("Excel.Application")
            wkbk = xlApp.Workbooks.open(bomfile)
            wksht = wkbk.Worksheets(1)
            col = 5
            count = 0
            missing = []
            row = int(wksht.UsedRange.Rows.Count) + 3
            wksht.Cells(row, 2).Value = "Missing values in Bottom pcb:"
            for item in bot_bom_data:
                if item[2] == 0:
                    if top_bom_data[count][2] == 0:
                        missing.append(item[0])
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col+1).Value = item[1]
                    row += 1
                count += 1

            count = 0
            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in Top pcb:"
            for item in top_bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col+1).Value = item[1]
                    row += 1
                count += 1

            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in both:"
            for item in missing:
                if bool(item):
                    val = ", ".join(item)
                    wksht.Cells(row, col).Value = val
                    row += 1

            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in BOM:"
            for item in botrefdes:
                if item[1] == "Not found":
                    # wksht.Cells(row, col).Font.ColorIndex = 3
                    wksht.Cells(row, col).Value = item[0]
                    row += 1
            for item in toprefdes:
                if item[1] == "Not found":
                    # wksht.Cells(row, col).Font.ColorIndex = 3
                    wksht.Cells(row, col).Value = item[0]
                    row += 1

            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()

        else:
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            col = 5
            count = 0
            missing = []
            row = int(ws_bom.max_row) + 3
            ws_bom.cell(row=row, column=2).value = "Missing values in Bottom pcb:"
            for item in bot_bom_data:
                if item[2] == 0:
                    if top_bom_data[count][2] == 0:
                        missing.append(item[0])
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col+1).value = item[1]
                    row += 1
                count += 1

            count = 0
            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in Top pcb:"
            for item in top_bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col+1).value = item[1]
                    row += 1
                count += 1

            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in both:"
            for item in missing:
                if bool(item):
                    val = ", ".join(item)
                    ws_bom.cell(row=row, column=col).value = val
                    row += 1

            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in BOM:"
            for item in botrefdes:
                if item[1] == "Not found":
                    # ws_bom.cell(row=row, column=col).font = Font(color="00FF0000")
                    ws_bom.cell(row=row, column=col).value = item[0]
                    row += 1
            for item in toprefdes:
                if item[1] == "Not found":
                    # ws_bom.cell(row=row, column=col).font = Font(color="00FF0000")
                    ws_bom.cell(row=row, column=col).value = item[0]
                    row += 1

            wb_bom.save(bomfile)
        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 309")
        logging.error("Error while writing to BOM excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing to BOM excel file!")
        if file_extension == ".xls" or file_extension == ".XLS":
            if bool(wkbk):
                wkbk.Close()
                xlApp.Quit()
        else:
            if bool(wb_bom):
                wb_bom.close()
        sys.exit(1)


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")

    Bom_File, bot_file, top_file = getfiles()
    pre_botrefdes, pre_toprefdes = read_pcb(bot_file, top_file)
    bom_dict = BOM_metadata()
    pre_bot_bom_data, pre_top_bom_data = readBom(Bom_File, bom_dict)
    post_botrefdes, post_toprefdes, post_bot_bom_data, post_top_bom_data = mapping(pre_botrefdes, pre_toprefdes, pre_bot_bom_data, pre_top_bom_data)
    Write_PCB(bot_file, top_file, post_botrefdes, post_toprefdes)
    Write_Bom(Bom_File, post_bot_bom_data, post_top_bom_data, post_botrefdes, post_toprefdes)

    logging.info("Successfully Executed!!!\n\n")
    print("Successfully Executed!!!")
    messagebox.showinfo(title="Status", message="Completed!!!")
