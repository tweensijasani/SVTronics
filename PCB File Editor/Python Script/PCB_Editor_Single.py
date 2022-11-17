import re
import sys
import xlrd
import pathlib
import logging
import openpyxl
import tkinter as tk
from easygui import *
from win32com import client
from openpyxl.styles import Font
from tkinter import filedialog, messagebox

logging.basicConfig(level=logging.DEBUG, filename="PCB_single_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")

root = tk.Tk()
root.withdraw()


def getfiles():
    try:
        counter = 0
        pcb_file = filedialog.askopenfilename(title="Select PCB File")
        while pcb_file == '' and counter < 1:
            messagebox.showerror(title="File Error", message="PCB File Not Selected!")
            logging.warning("PCB File Not Selected!")
            pcb_file = filedialog.askopenfilename(title="Select PCB File")
            counter += 1
        if pcb_file == '':
            messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
            logging.error("Terminated: PCB File not selected!")
            print("Terminated: PCB File not selected!")
            sys.exit(1)
        else:
            logging.info("PCB File Selected!")

        counter = 0
        Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                              filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
        while Bom_File == '' and counter < 1:
            messagebox.showerror(title="File Error", message="BOM File Not Selected!")
            logging.warning("BOM File Not Selected!")
            Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                                  filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
            counter += 1
        if Bom_File == '':
            messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
            logging.error("Terminated: Customer BOM not selected!")
            print("Terminated: Customer BOM not selected!")
            sys.exit(1)
        else:
            logging.info("Customer BOM Excel Selected!")
        modify(Bom_File, pcb_file)

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__}")
        logging.error("Error while fetching files!")
        logging.error(f"{e}")
        print(e, "\n Error while fetching files!")
        sys.exit(1)


def modify(bomfile, pcb_file):

    try:
        logging.info("Reading .pcb file...")
        pcbrefdes = []
        pcb_textfile = open(pcb_file, 'r')
        pcbfiledata = pcb_textfile.readlines()
        for line in pcbfiledata:
            match = re.match("^F9\s", line)
            if match:
                pcbrefdes.append([line.strip().split(" ").pop()])
        logging.info("Finished reading pcb file!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 66")
        logging.error("Error while reading .pcb file!")
        logging.error(f"{e}")
        print(e, "\n Error while reading .pcb file!")
        sys.exit(1)

    try:
        logging.info("Getting BOM column info...")
        text = "BOM File Details"
        title = "Enter Details"
        input_list = ["RefDes Column", "Manex P/N Column", "Start Row", "End Row", "Delimiter", "Separator"]
        output = multenterbox(text, title, input_list)

        while output[0] is None or output[1] is None or output[2] is None or output[3] is None or not output[0].isalpha() \
                or not output[1].isalpha() or not output[2].isnumeric() or not output[3].isnumeric():
            messagebox.showerror(title="Invalid Format", message="Please enter valid text formats")
            text = "BOM File Details"
            title = "Enter Details"
            input_list = ["RefDes Column", "Manex P/N Column", "Start Row", "End Row", "Delimiter", "Separator"]
            output = multenterbox(text, title, input_list)

        bom_designator = (output[0].strip()).upper()
        bom_pn = (output[1].strip()).upper()
        bom_start_row = int(output[2].strip())
        bom_end_row = int(output[3].strip())
        bom_delimiter = output[4].strip()
        bom_separator = output[5].strip()

        logging.info("Reading BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls":
            wb_bom = xlrd.open_workbook(bomfile)
            ws_bom = wb_bom.sheet_by_index(0)
            bom_data = []
            bom_col_des = ord(bom_designator) - 65
            bom_col_pn = ord(bom_pn) - 65
            for row in range(bom_start_row - 1, bom_end_row):
                var = ws_bom.row_values(row)
                x = var[bom_col_des]
                if bom_delimiter != '':
                    x = (var[bom_col_des]).replace(" ", "").split(bom_delimiter)
                    x = list(filter(None, x))
                    if bom_separator != '':
                        for item in x:
                            if bom_separator in item:
                                res = []
                                stry = item.split(bom_separator)
                                str1 = stry[0]
                                str2 = stry[1]
                                base = ""
                                for i in range(len(str1) - 1):
                                    if str1[i] == str2[i]:
                                        base = f"{base}{str1[i]}"
                                    else:
                                        break
                                str1 = str1.lstrip(base)
                                str2 = str2.lstrip(base)
                                my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                                my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                                if len(my_list1) > 1:
                                    if my_list1[0].isalpha():
                                        for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                res.append(f"{base}{chr(i)}{j}")
                                    else:
                                        for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                            for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{i}{chr(j)}")
                                else:
                                    if my_list1[0].isalpha():
                                        for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            res.append(f"{base}{chr(i)}")
                                    else:
                                        for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                            res.append(f"{base}{j}")
                                pointer = x.index(item)
                                x.pop(pointer)
                                x.extend(res)
                bom_data.append([x, var[bom_col_pn], 0])
        else:
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            bom_data = []
            bom_col_des = ord(bom_designator) - 65
            bom_col_pn = ord(bom_pn) - 65
            for row in bom_rows[int(bom_start_row) - 1:int(bom_end_row)]:
                x = row[bom_col_des].value
                if bom_delimiter != '':
                    x = (row[bom_col_des].value).replace(" ", "").split(bom_delimiter)
                    x = list(filter(None, x))
                    if bom_separator != '':
                        for item in x:
                            if bom_separator in item:
                                res = []
                                stry = item.split(bom_separator)
                                str1 = stry[0]
                                str2 = stry[1]
                                base = ""
                                for i in range(len(str1) - 1):
                                    if str1[i] == str2[i]:
                                        base = f"{base}{str1[i]}"
                                    else:
                                        break
                                str1 = str1.lstrip(base)
                                str2 = str2.lstrip(base)
                                my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                                my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                                if len(my_list1) > 1:
                                    if my_list1[0].isalpha():
                                        for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                res.append(f"{base}{chr(i)}{j}")
                                    else:
                                        for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                            for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{i}{chr(j)}")
                                else:
                                    if my_list1[0].isalpha():
                                        for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                            res.append(f"{base}{chr(i)}")
                                    else:
                                        for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                            res.append(f"{base}{j}")
                                pointer = x.index(item)
                                x.pop(pointer)
                                x.extend(res)
                bom_data.append([x, row[bom_col_pn].value, 0])
        logging.info("Finished reading BOM excel!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 105")
        logging.error("Error while reading BOM Excel!")
        logging.error(f"{e}")
        print(e, "\n Error while reading BOM Excel!")
        sys.exit(1)

    try:
        logging.info("Mapping PCB RefDes to BOM excel...")
        pointer = 0
        for value in pcbrefdes:
            flag = 0
            for ref in bom_data:
                if value[0] in ref[0]:
                    pcbrefdes[pointer].append(ref[1])
                    ref[2] = 1
                    flag = 1
                    break
            if flag == 0:
                pcbrefdes[pointer].append("Not found")
            pointer += 1
        logging.info("Finished mapping!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 165")
        logging.error("Error while mapping designators from .pcb file to excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from .pcb file to excel file!")
        sys.exit(1)

    try:
        logging.info("Editing pcb content...")
        pointer = 0
        line_pointer = 0
        for line in pcbfiledata:
            match = re.match("^F8\s", line)
            if match:
                if pcbrefdes[pointer][1] != "Not found":
                    x = line.strip().split(" ")
                    x.pop()
                    x.append(pcbrefdes[pointer][1])
                    x = " ".join(x)
                    pcbfiledata[line_pointer] = x + "\n"
                    pointer += 1
            line_pointer += 1
        logging.info("Finished editing!")

        pcb_textfile.close()

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 190")
        logging.error("Error while editing .pcb file content!")
        logging.error(f"{e}")
        print(e, "\n Error while editing .pcb file content!")
        sys.exit(1)

    try:
        logging.info("Writing modified pcb file...")
        var = pcb_textfile.name.split("/")
        new_file_name = var.pop().replace(".pcb", "_modified.pcb")
        var.append(new_file_name)
        new_file = "/".join(var)

        with open(new_file, "w") as f:
            for item in pcbfiledata:
                f.write("%s" % item)
        f.close()
        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 215")
        logging.error("Error while creating modified .pcb file!")
        logging.error(f"{e}")
        print(e, "\n Error while creating modified .pcb file!")
        sys.exit(1)

    try:
        logging.info("Writing to BOM excel...")
        if file_extension == ".xls":
            xlApp = client.Dispatch("Excel.Application")
            wkbk = xlApp.Workbooks.open(bomfile)
            wksht = wkbk.Worksheets(1)
            col = 5
            count = 0
            missing = []
            row = int(wksht.UsedRange.Rows.Count) + 3
            wksht.Cells(row, 2).Value = "Missing values in PCB File:"
            for item in bom_data:
                if item[2] == 0:
                    val = ", ".join(item[0])
                    wksht.Cells(row, col).Value = val
                    row += 1
                count += 1

            row = row + 2
            wksht.Cells(row, 2).Value = "Missing values in BOM:"
            for item in pcbrefdes:
                if item[1] == "Not found":
                    wksht.Cells(row, col).Font.ColorIndex = 3
                    wksht.Cells(row, col).Value = item[0]
                    row += 1

            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()

        else:
            col = 5
            count = 0
            row = int(ws_bom.max_row) + 3
            ws_bom.cell(row=row, column=2).value = "Missing values in PCB File:"
            for item in bom_data:
                if item[2] == 0:
                    val = ", ".join(item[0])
                    ws_bom.cell(row=row, column=col).value = val
                    row += 1
                count += 1

            row = row + 2
            ws_bom.cell(row=row, column=2).value = "Missing values in BOM:"
            for item in pcbrefdes:
                if item[1] == "Not found":
                    ws_bom.cell(row=row, column=col).font = Font(color="00FF0000")
                    ws_bom.cell(row=row, column=col).value = item[0]
                    row += 1
            wb_bom.save(bomfile)
        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 239")
        logging.error("Error while writing to BOM excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing to BOM excel file!")
        sys.exit(1)


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")
    getfiles()
    logging.info("Sucessfully Executed!!!\n\n")
    print("Sucessfully Executed!!!")
    messagebox.showinfo(title="Status", message="Completed!!!")
