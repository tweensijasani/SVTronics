import re
import sys
import xlrd
import pathlib
import logging
import openpyxl
import easygui.boxes
import tkinter as tk
from easygui import *
from win32com import client
from tkinter import filedialog, messagebox


logging.basicConfig(level=logging.DEBUG, filename="mmd_logfile.txt", filemode="a+",
                    format="%(asctime)-15s %(levelname)-8s %(message)s")

root = tk.Tk()
root.withdraw()

easygui.boxes.global_state.PROPORTIONAL_FONT_FAMILY = ("MS", "Arial")
easygui.boxes.global_state.MONOSPACE_FONT_FAMILY = "Arial"
easygui.boxes.global_state.PROPORTIONAL_FONT_SIZE = 14
easygui.boxes.global_state.MONOSPACE_FONT_SIZE = 14
easygui.boxes.global_state.TEXT_ENTRY_FONT_SIZE = 14
easygui.boxes.global_state.prop_font_line_length = 30
easygui.boxes.global_state.fixw_font_line_length = 40


def getfiles():

    try:
        file_count = tk.messagebox.askyesno(title="MMD Files", message="Do you have separate TOP-BOT mmd files?")
        if file_count is True:
            counter = 0
            bot_file = filedialog.askopenfilename(title="Select Bottom MMD File")
            while bot_file == '' and counter < 1:
                messagebox.showerror(title="File Error", message="MMD BOT File Not Selected!")
                logging.warning("MMD BOT File Not Selected!")
                bot_file = filedialog.askopenfilename(title="Select Bottom MMD File")
                counter += 1
            if bot_file == '':
                messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
                logging.error("Terminated: MMD BOT not selected!")
                print("Terminated: MMD BOT not selected!")
                sys.exit(1)
            else:
                logging.info("MMD BOT File Selected!")

            counter = 0
            top_file = filedialog.askopenfilename(title="Select Top MMD File")
            while top_file == '' and counter < 1:
                messagebox.showerror(title="File Error", message="MMD TOP File Not Selected!")
                logging.warning("MMD TOP File Not Selected!")
                top_file = filedialog.askopenfilename(title="Select Top MMD File")
                counter += 1
            if top_file == '':
                messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
                logging.error("Terminated: MMD TOP not selected!")
                print("Terminated: MMD TOP not selected!")
                sys.exit(1)
            else:
                logging.info("MMD TOP File Selected!")

            counter = 0
            Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                                  filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
            while Bom_File == '' and counter < 1:
                messagebox.showerror(title="File Error", message="BOM File Not Selected!")
                logging.warning("BOM File Not Selected!")
                Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                                      filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
                counter += 1
            if Bom_File == '':
                messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
                logging.error("Terminated: Customer BOM not selected!")
                print("Terminated: Customer BOM not selected!")
                sys.exit(1)
            else:
                logging.info("Customer BOM Excel Selected!")

            return Bom_File, bot_file, top_file

        else:
            counter = 0
            mmd_file = filedialog.askopenfilename(title="Select MMD File")
            while mmd_file == '' and counter < 1:
                messagebox.showerror(title="File Error", message="MMD File Not Selected!")
                logging.warning("MMD File Not Selected!")
                mmd_file = filedialog.askopenfilename(title="Select MMD File")
                counter += 1
            if mmd_file == '':
                messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
                logging.error("Terminated: MMD File not selected!")
                print("Terminated: MMD File not selected!")
                sys.exit(1)
            else:
                logging.info("MMD File Selected!")

            counter = 0
            Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                                  filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
            while Bom_File == '' and counter < 1:
                messagebox.showerror(title="File Error", message="BOM File Not Selected!")
                logging.warning("BOM File Not Selected!")
                Bom_File = filedialog.askopenfilename(title="Select BOM file",
                                                      filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xls")))
                counter += 1
            if Bom_File == '':
                messagebox.showerror(title="Invalid Input", message="Something went wrong!! Please try again....")
                logging.error("Terminated: Customer BOM not selected!")
                print("Terminated: Customer BOM not selected!")
                sys.exit(1)
            else:
                logging.info("Customer BOM Excel Selected!")

            return Bom_File, mmd_file, False

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__}")
        logging.error("Error while fetching files!")
        logging.error(f"{e}")
        print(e, "\n Error while fetching files!")
        sys.exit(1)


def read_both_mmd(bot_file, top_file):
    try:
        logging.info("Reading bot.mmd file...")
        bot_textfile = open(bot_file, 'r')
        botfiledata = bot_textfile.readlines()
        bot_list = []
        for line in botfiledata:
            bot_list.append([line])

        logging.info("Reading top.mmd file...")
        top_textfile = open(top_file, 'r')
        topfiledata = top_textfile.readlines()
        top_list = []
        for line in topfiledata:
            top_list.append([line])

        bot_textfile.close()
        top_textfile.close()
        logging.info("Finished reading mmd files!")
        return bot_list, top_list

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 80")
        logging.error("Error while reading .mmd files!")
        logging.error(f"{e}")
        print(e, "\n Error while reading .mmd files!")
        sys.exit(1)


def BOM_metadata():
    try:
        logging.info("Getting BOM column info...")
        text = "BOM File Details"
        title = "Enter Details"
        input_list = ["RefDes Column", "Manex P/N Column", "Start Row", "End Row", "Delimiter", "Separator"]
        output = multenterbox(text, title, input_list)

        while output[0] is None or output[1] is None or output[2] is None or output[3] is None or not output[0].isalpha() \
                or not output[1].isalpha() or not output[2].isnumeric() or not output[3].isnumeric():
            messagebox.showerror(title="Invalid Format", message="Please enter valid text formats")
            text = "BOM File Details"
            title = "Enter Details"
            input_list = ["RefDes Column", "Manex P/N Column", "Start Row", "End Row", "Delimiter", "Separator"]
            output = multenterbox(text, title, input_list)

        bom_designator = (output[0].strip()).upper()
        bom_pn = (output[1].strip()).upper()
        bom_start_row = int(output[2].strip())
        bom_end_row = int(output[3].strip())
        bom_delimiter = output[4].strip()
        bom_separator = output[5].strip()

        bom_dict = {"bom_designator": bom_designator, "bom_pn": bom_pn, "bom_start_row": bom_start_row,
                    "bom_end_row": bom_end_row,
                    "bom_delimiter": bom_delimiter, "bom_separator": bom_separator}
        logging.info("Info populated!")
        return bom_dict

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 130")
        logging.error("Error while getting BOM metadata!")
        logging.error(f"{e}")
        print(e, "\n Error while getting BOM metadata!")
        sys.exit(1)


def readBOM(bomfile, bom_dict):
    try:
        messagebox.showinfo(title="Permission",
                            message="Please close the BOM file if open. \nHit OK only when the file is closed!")

        logging.info("Reading BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            wb_bom = xlrd.open_workbook(bomfile)
            ws_bom = wb_bom.sheet_by_index(0)
            bot_bom_data = []
            top_bom_data = []
            bom_col_des = ord(bom_dict["bom_designator"]) - 65
            bom_col_pn = ord(bom_dict["bom_pn"]) - 65
            for row in range(bom_dict["bom_start_row"] - 1, bom_dict["bom_end_row"]):
                var = ws_bom.row_values(row)
                x = var[bom_col_des]
                if bool(x):
                    if bom_dict["bom_delimiter"] != '':
                        x = (var[bom_col_des]).replace(" ", "").split(bom_dict["bom_delimiter"])
                        x = list(filter(None, x))
                        if bom_dict["bom_separator"] != '':
                            for item in x:
                                if bom_dict["bom_separator"] in item:
                                    res = []
                                    stry = item.split(bom_dict["bom_separator"])
                                    str1 = stry[0]
                                    str2 = stry[1]
                                    base = ""
                                    for i in range(len(str1) - 1):
                                        if str1[i] == str2[i]:
                                            base = f"{base}{str1[i]}"
                                        else:
                                            break
                                    if base == '':
                                        break
                                    count1 = 0
                                    count2 = 0
                                    for i in range(len(base) - 1):
                                        if str1[i] == base[i]:
                                            count1 += 1
                                        if str2[i] == base[i]:
                                            count2 += 1
                                    str1 = str1[count1 + 1:]
                                    str2 = str2[count2 + 1:]
                                    my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                                    my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                                    if len(my_list1) > 1:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                    res.append(f"{base}{chr(i)}{j}")
                                        else:
                                            for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                    res.append(f"{base}{i}{chr(j)}")
                                    else:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{chr(i)}")
                                        else:
                                            for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                                res.append(f"{base}{j}")
                                    pointer = x.index(item)
                                    x.pop(pointer)
                                    x.extend(res)
                y = str(var[bom_col_pn])
                if bool(y):
                    y = y.split(".")
                    bot_bom_data.append([x, y[0], 0])
                    top_bom_data.append([x, y[0], 0])
                else:
                    bot_bom_data.append([x, y, 0])
                    top_bom_data.append([x, y, 0])
        else:
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            bom_rows = list(ws_bom.rows)
            bot_bom_data = []
            top_bom_data = []
            bom_col_des = ord(bom_dict["bom_designator"]) - 65
            bom_col_pn = ord(bom_dict["bom_pn"]) - 65
            for row in bom_rows[int(bom_dict["bom_start_row"]) - 1:int(bom_dict["bom_end_row"])]:
                x = row[bom_col_des].value
                if bool(x):
                    if bom_dict["bom_delimiter"] != '':
                        x = (row[bom_col_des].value).replace(" ", "").split(bom_dict["bom_delimiter"])
                        x = list(filter(None, x))
                        if bom_dict["bom_separator"] != '':
                            for item in x:
                                if bom_dict["bom_separator"] in item:
                                    res = []
                                    stry = item.split(bom_dict["bom_separator"])
                                    str1 = stry[0]
                                    str2 = stry[1]
                                    base = ""
                                    for i in range(len(str1) - 1):
                                        if str1[i] == str2[i]:
                                            base = f"{base}{str1[i]}"
                                        else:
                                            break
                                    if base == '':
                                        break
                                    count1 = 0
                                    count2 = 0
                                    for i in range(len(base) - 1):
                                        if str1[i] == base[i]:
                                            count1 += 1
                                        if str2[i] == base[i]:
                                            count2 += 1
                                    str1 = str1[count1 + 1:]
                                    str2 = str2[count2 + 1:]
                                    my_list1 = list(filter(None, re.split(r'(\d+)', str1)))
                                    my_list2 = list(filter(None, re.split(r'(\d+)', str2)))
                                    if len(my_list1) > 1:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                for j in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                    res.append(f"{base}{chr(i)}{j}")
                                        else:
                                            for i in range(int(my_list1[1]), int(my_list2[1]) + 1):
                                                for j in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                    res.append(f"{base}{i}{chr(j)}")
                                    else:
                                        if my_list1[0].isalpha():
                                            for i in range(ord(my_list1[0]), ord(my_list2[0]) + 1):
                                                res.append(f"{base}{chr(i)}")
                                        else:
                                            for j in range(int(my_list1[0]), int(my_list2[0]) + 1):
                                                res.append(f"{base}{j}")
                                    pointer = x.index(item)
                                    x.pop(pointer)
                                    x.extend(res)
                y = str(row[bom_col_pn].value)
                if bool(y):
                    y = y.split(".")
                    bot_bom_data.append([x, y[0], 0])
                    top_bom_data.append([x, y[0], 0])
                else:
                    bot_bom_data.append([x, y, 0])
                    top_bom_data.append([x, y, 0])
            wb_bom.close()
        logging.info("Finished reading BOM excel!")
        return bot_bom_data, top_bom_data

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 130")
        logging.error("Error while reading BOM Excel!")
        logging.error(f"{e}")
        print(e, "\n Error while reading BOM Excel!")
        sys.exit(1)


def Map_both(bot_bom_data, top_bom_data, bot_list, top_list):
    try:
        logging.info("Mapping bot.mmd RefDes to BOM excel...")
        pointer = 0
        missing_bot = []
        for value in bot_list:
            if re.match("^#", value[0]):
                match = value[0].strip().split("\t")
                refdes = match.pop()
                flag = 0
                for ref in bot_bom_data:
                    if bool(ref[0]) and refdes in ref[0] and bool(ref[1]):
                        match = match[0:3]
                        match.append(ref[1])
                        ref[2] = 1
                        flag = 1
                        break
                if flag == 0:
                    missing_bot.append(refdes)
                match.append(refdes)
                match = "\t".join(match)
                bot_list[pointer][0] = match + "\n"
            pointer += 1

        logging.info("Mapping top.mmd RefDes to BOM excel...")
        pointer = 0
        missing_top = []
        for value in top_list:
            if re.match("^#", value[0]):
                match = value[0].strip().split("\t")
                refdes = match.pop()
                flag = 0
                for ref in top_bom_data:
                    if bool(ref[0]) and refdes in ref[0] and bool(ref[1]):
                        match = match[0:3]
                        match.append(ref[1])
                        ref[2] = 1
                        flag = 1
                        break
                if flag == 0:
                    missing_top.append(refdes)
                match.append(refdes)
                match = "\t".join(match)
                top_list[pointer][0] = match + "\n"
            pointer += 1

        logging.info("Finished mapping!")
        return bot_bom_data, top_bom_data, bot_list, top_list, missing_top, missing_bot

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 192")
        logging.error("Error while mapping designators from .mmd files to excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from .mmd files to excel file!")
        sys.exit(1)


def Write_both_mmd(bot_textfile, top_textfile, bot_list, top_list):
    try:
        logging.info("Writing modified bot.mmd file...")
        var = bot_textfile.name.split("/")
        last_item = var.pop()
        if ".mmd" in last_item:
            new_file_name = last_item.replace(".mmd", "_Svt_PartNo.mmd")
        else:
            new_file_name = last_item.replace(".MMD", "_Svt_PartNo.MMD")
        var.append(new_file_name)
        new_file = "/".join(var)

        with open(new_file, "w") as f:
            for item in bot_list:
                f.write("%s" % item[0])
        f.close()

        logging.info("Writing modified top.mmd file...")
        var = top_textfile.name.split("/")
        last_item = var.pop()
        if ".mmd" in last_item:
            new_file_name = last_item.replace(".mmd", "_Svt_PartNo.mmd")
        else:
            new_file_name = last_item.replace(".MMD", "_Svt_PartNo.MMD")
        var.append(new_file_name)
        new_file = "/".join(var)

        with open(new_file, "w") as f:
            for item in top_list:
                f.write("%s" % item[0])
        f.close()

        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 272")
        logging.error("Error while creating modified .mmd files!")
        logging.error(f"{e}")
        print(e, "\n Error while creating modified .mmd files!")
        sys.exit(1)


def Write_BOM(bomfile, bot_bom_data, top_bom_data, missing_bot, missing_top, bot_file, top_file):
    try:
        logging.info("Writing to BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            xlApp = client.Dispatch("Excel.Application")
            wkbk = xlApp.Workbooks.open(bomfile)
            wksht = wkbk.Worksheets(1)
            col = 5
            count = 0
            missing = []
            row = int(wksht.UsedRange.Rows.Count) + 3
            wksht.Cells(row, 2).Value = f"Missing values in '{bot_file}':"
            for item in bot_bom_data:
                if item[2] == 0:
                    if top_bom_data[count][2] == 0:
                        missing.append([item[0], item[1]])
                    val = ", ".join(item[0])
                    wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col+1).Value = item[1]
                    row += 1
                count += 1

            count = 0
            row = row + 2
            wksht.Cells(row, 2).Value = f"Missing values in '{top_file}':"
            for item in top_bom_data:
                if item[2] == 0:
                    val = ", ".join(item[0])
                    wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col+1).Value = item[1]
                    row += 1
                count += 1

            if bool(missing):
                row = row + 2
                wksht.Cells(row, 2).Value = "Missing values in both:"
                for item in missing:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col + 1).Value = item[1]
                    row += 1

            if bool(missing_bot) or bool(missing_top):
                row = row + 2
                wksht.Cells(row, 2).Value = "Missing values in BOM:"
                for item in missing_bot:
                    wksht.Cells(row, col).Value = item
                    row += 1
                for item in missing_top:
                    wksht.Cells(row, col).Value = item
                    row += 1

            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()

        else:
            col = 5
            count = 0
            missing = []
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            row = int(ws_bom.max_row) + 3
            # ws_bom.cell(row=row, column=2).value = f"Missing values in '{bot_file}':"
            for item in bot_bom_data:
                if item[2] == 0:
                    if top_bom_data[count][2] == 0:
                        missing.append([item[0], item[1]])
                count += 1

            if bool(missing):
                # row = row + 2
                ws_bom.cell(row=row, column=2).value = "Missing values in both:"
                for item in missing:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col + 1).value = item[1]
                    row += 1

            if bool(missing_bot) or bool(missing_top):
                row = row + 2
                ws_bom.cell(row=row, column=2).value = "Missing values in BOM:"
                for item in missing_bot:
                    ws_bom.cell(row=row, column=col).value = item
                    row += 1
                for item in missing_top:
                    ws_bom.cell(row=row, column=col).value = item
                    row += 1

            wb_bom.save(bomfile)

        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 309")
        logging.error("Error while writing to BOM excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing to BOM excel file!")
        if file_extension == ".xls" or file_extension == ".XLS":
            if bool(wkbk):
                wkbk.Close()
                xlApp.Quit()
        else:
            if bool(wb_bom):
                wb_bom.close()
        sys.exit(1)


def read_mmd(mmdfile):
    try:
        logging.info("Reading mmd file...")
        mmd_textfile = open(mmdfile, 'r')
        mmdfiledata = mmd_textfile.readlines()
        mmd_list = []
        for line in mmdfiledata:
            mmd_list.append([line])

        mmd_textfile.close()
        logging.info("Finished reading mmd file!")
        return mmd_list

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 80")
        logging.error("Error while reading .mmd file!")
        logging.error(f"{e}")
        print(e, "\n Error while reading .mmd file!")
        sys.exit(1)


def Mapping(bom_data, mmd_list):
    try:
        logging.info("Mapping mmd RefDes to BOM excel...")
        pointer = 0
        missing = []
        for value in mmd_list:
            if re.match("^#", value[0]):
                match = value[0].strip().split("\t")
                refdes = match.pop()
                flag = 0
                for ref in bom_data:
                    if bool(ref[0]) and refdes in ref[0] and bool(ref[1]):
                        match = match[0:3]
                        match.append(ref[1])
                        ref[2] = 1
                        flag = 1
                        break
                if flag == 0:
                    missing.append(refdes)
                match.append(refdes)
                match = "\t".join(match)
                mmd_list[pointer][0] = match + "\n"
            pointer += 1

        logging.info("Finished mapping!")
        return bom_data, mmd_list, missing

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 192")
        logging.error("Error while mapping designators from .mmd file to excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while mapping designators from .mmd file to excel file!")
        sys.exit(1)


def Write_mmd(mmd_textfile, mmd_list):
    try:
        logging.info("Writing modified mmd file...")
        var = mmd_textfile.name.split("/")
        last_item = var.pop()
        if ".mmd" in last_item:
            new_file_name = last_item.replace(".mmd", "_Svt_PartNo.mmd")
        else:
            new_file_name = last_item.replace(".MMD", "_Svt_PartNo.MMD")
        var.append(new_file_name)
        new_file = "/".join(var)

        with open(new_file, "w") as f:
            for item in mmd_list:
                f.write("%s" % item[0])
        f.close()

        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 272")
        logging.error("Error while creating modified .mmd file!")
        logging.error(f"{e}")
        print(e, "\n Error while creating modified .mmd file!")
        sys.exit(1)


def Write_Single_BOM(bomfile, bom_data, missing, file):
    try:
        logging.info("Writing to BOM excel...")
        file_extension = pathlib.Path(bomfile).suffix
        if file_extension == ".xls" or file_extension == ".XLS":
            xlApp = client.Dispatch("Excel.Application")
            wkbk = xlApp.Workbooks.open(bomfile)
            wksht = wkbk.Worksheets(1)
            col = 5
            count = 0
            row = int(wksht.UsedRange.Rows.Count) + 3
            wksht.Cells(row, 2).Value = f"Missing values in '{file}':"
            for item in bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        wksht.Cells(row, col).Value = val
                    wksht.Cells(row, col+1).Value = item[1]
                    row += 1
                count += 1

            if bool(missing):
                row = row + 2
                wksht.Cells(row, 2).Value = "Missing values in BOM:"
                for item in missing:
                    wksht.Cells(row, col).Value = item
                    row += 1

            wkbk.Save()
            wkbk.Close(True)
            xlApp.Quit()

        else:
            col = 5
            count = 0
            wb_bom = openpyxl.load_workbook(bomfile)
            ws_bom = wb_bom.worksheets[0]
            row = int(ws_bom.max_row) + 3
            ws_bom.cell(row=row, column=2).value = f"Missing values in '{file}':"
            for item in bom_data:
                if item[2] == 0:
                    if bool(item[0]):
                        val = ", ".join(item[0])
                        ws_bom.cell(row=row, column=col).value = val
                    ws_bom.cell(row=row, column=col+1).value = item[1]
                    row += 1
                count += 1

            if bool(missing):
                row = row + 2
                ws_bom.cell(row=row, column=2).value = "Missing values in BOM:"
                for item in missing:
                    ws_bom.cell(row=row, column=col).value = item
                    row += 1

            wb_bom.save(bomfile)

        logging.info("Finished writing!")

    except Exception as e:
        messagebox.showerror(title=f"{e.__class__}", message="Something went wrong!! Please try again....")
        logging.error(f"{e.__class__} from line 309")
        logging.error("Error while writing to BOM excel file!")
        logging.error(f"{e}")
        print(e, "\n Error while writing to BOM excel file!")
        if file_extension == ".xls" or file_extension == ".XLS":
            if bool(wkbk):
                wkbk.Close()
                xlApp.Quit()
        else:
            if bool(wb_bom):
                wb_bom.close()
        sys.exit(1)


if __name__ == "__main__":
    logging.info("Execution Started...")
    print("Execution Started!!!")

    Bom_File, bot_file, top_file = getfiles()
    bom_dict = BOM_metadata()
    pre_bot_bom_data, pre_top_bom_data = readBOM(Bom_File, bom_dict)
    if bool(top_file):
        pre_bot_list, pre_top_list = read_both_mmd(bot_file, top_file)
        post_bot_bom_data, post_top_bom_data, post_bot_list, post_top_list, missing_top, missing_bot = Map_both(pre_bot_bom_data, pre_top_bom_data, pre_bot_list, pre_top_list)
        Write_both_mmd(bot_file, top_file, post_bot_list, post_top_list)
        Write_BOM(Bom_File, post_bot_bom_data, post_top_bom_data, missing_bot, missing_top, bot_file, top_file)
    else:
        pre_mmd_list = read_mmd(bot_file)
        post_bom_data, post_mmd_list, missing = Mapping(pre_bot_bom_data, pre_mmd_list)
        Write_mmd(bot_file, post_mmd_list)
        Write_Single_BOM(Bom_File, post_bom_data, missing, bot_file)

    logging.info("Successfully Executed!!!\n\n")
    print("Successfully Executed!!!")
    messagebox.showinfo(title="Status", message="Completed!!!")
